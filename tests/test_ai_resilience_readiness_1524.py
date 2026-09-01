from __future__ import annotations

import json
from io import BytesIO

from openpyxl import load_workbook

from core.ai_gateway import AIResult, FailoverProvider, _extract_json
from core.atomic_verification_engine import verify_checklist_rows
from core.semantic_evidence_engine import _call_batches, _preflight_provider
from studio.data import structured_excel_report


class _StaticProvider:
    def __init__(self, name: str, result: AIResult):
        self.name = name
        self.api_key = "configured"
        self.result = result
        self.calls = 0

    def generate(self, prompt: str, system: str = "") -> AIResult:
        self.calls += 1
        return self.result


class _SplitProvider:
    name = "SplitProvider"
    api_key = "configured"

    def __init__(self):
        self.batch_sizes: list[int] = []

    def generate(self, prompt: str, system: str = "") -> AIResult:
        payload = json.loads(prompt)
        packets = payload["packets"]
        self.batch_sizes.append(len(packets))
        if len(packets) > 1:
            return AIResult(True, self.name, text="не JSON", status_code=200, model="split-test")
        packet_id = packets[0]["packet_id"]
        return AIResult(True, self.name, text=json.dumps({
            "decisions": [{"packet_id": packet_id, "verdict": "INSUFFICIENT"}],
        }), status_code=200, model="split-test")


def test_json_recovery_preserves_closed_rows_from_truncated_root_array():
    parsed = _extract_json(
        'Ответ:\n```json\n{"decisions":['
        '{"packet_id":"P-1","verdict":"INSUFFICIENT"},'
        '{"packet_id":"P-2","verdict":"SUPPORTS"},'
    )
    assert isinstance(parsed, dict)
    assert [row["packet_id"] for row in parsed["decisions"]] == ["P-1", "P-2"]


def test_failover_isolates_forbidden_provider_and_uses_next_lane():
    blocked = _StaticProvider("OpenRouter", AIResult(False, "OpenRouter", error="forbidden", status_code=403))
    good = _StaticProvider("Groq", AIResult(True, "Groq", text='{"decisions":[]}', status_code=200, model="test"))
    result = FailoverProvider([blocked, good]).generate_validated(
        "{}", "Верните JSON", lambda text: isinstance(_extract_json(text).get("decisions"), list),
    )
    assert result.ok and result.provider == "Groq"
    assert blocked.calls == good.calls == 1


def test_batch_contract_failure_splits_to_single_packets_without_losing_queue():
    provider = _SplitProvider()
    packets = [{"packet_id": f"P-{index}"} for index in range(1, 5)]
    responses, errors, calls = _call_batches(provider, packets, batch_size=4, max_calls=20)
    assert set(responses) == {"P-1", "P-2", "P-3", "P-4"}
    assert provider.batch_sizes[0] == 4
    assert provider.batch_sizes.count(1) == 4
    assert any(row["state"] == "CONTRACT_RETRY_SPLIT" for row in calls)
    assert errors


def test_preflight_checks_working_json_contract_not_only_connectivity():
    provider = _StaticProvider("OpenRouter", AIResult(True, "OpenRouter", text="OK", status_code=200, model="free"))
    provider.test_connection = lambda: AIResult(True, "OpenRouter", text="OK", status_code=200, model="free")
    result = _preflight_provider(provider, "JUDGE")
    assert result["state"] == "FAILED"
    assert result["status_code"] == 422


def test_structured_parameter_presence_closes_safe_checklist_l2_check():
    checklist = [{
        "question": "Проверить наличие значения производительности",
        "automatic_section": "ТХ",
        "compiled_rule": {
            "typed_check": "ENGINEERING_PARAMETER_PRESENCE",
            "verification_level": "L2_VALUE",
            "parameter_codes": ["CAPACITY"],
        },
        "typed_check": "ENGINEERING_PARAMETER_PRESENCE",
    }]
    fact = {
        "property_code": "CAPACITY", "value": 1600, "unit": "тыс. т/год",
        "document": "ТХ.pdf", "page": 17, "document_type": "ТХ",
        "source_trace": "Производительность комплекса 1600 тыс. т/год",
        "fact_admission_decision": "ADMIT", "binding_status": "ROW_LOCKED",
    }
    result = verify_checklist_rows(
        checklist, knowledge_root="knowledge",
        fact_graph={"facts": [fact], "passages": []}, page_corpus=[],
    )
    assert checklist[0]["status"] == "Да"
    assert checklist[0]["verification_kind"] == "VERIFIED_OK"
    assert checklist[0]["evidence_level"] == "L5"
    assert result["summary"]["promoted_verified"] == 1


def test_addressable_drawing_title_closes_only_l1_presence_check():
    checklist = [{
        "question": "Проверить наличие плана организации рельефа",
        "automatic_section": "ПЗУ",
        "compiled_rule": {
            "typed_check": "DRAWING_PRESENCE_CHECK",
            "verification_level": "L1_PRESENCE",
            "evidence_terms": ["план", "организации", "рельефа"],
        },
        "typed_check": "DRAWING_PRESENCE_CHECK",
    }]
    pages = [{
        "document": "Раздел ПД №2_ПЗУ2.pdf", "document_type": "ПЗУ", "section": "ПЗУ", "page": 3,
        "source_modality": "DRAWING", "text": "План организации рельефа М 1:1000",
    }]
    result = verify_checklist_rows(
        checklist, knowledge_root="knowledge",
        fact_graph={"facts": [], "passages": pages}, page_corpus=pages,
    )
    assert checklist[0]["status"] == "Да"
    assert checklist[0]["verification_kind"] == "VERIFIED_OK"
    assert result["summary"]["promoted_verified"] == 1


def test_integrity_pass_does_not_claim_verification_is_complete():
    question = {
        "plan_id": "CHECK-1", "domain": "Чек-листы", "title": "Проверить схему",
        "verification_kind": "REVIEW_QUESTION", "status": "Требует проверки",
        "coverage_reason": "Нужна смысловая проверка", "evidence_level": "L4",
    }
    domains = {
        "Задание на проектирование": {"total": 0}, "НТД": {"total": 0},
        "Чек-листы": {"total": 1, "review": 1, "system_limitation": 0, "coverage_pct": 0},
    }
    docs = [{
        "Файл": "ИОС2.pdf", "Раздел": "ИОС2", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [question], "project_findings": 0, "review_questions": 1,
            "system_limitations": 0, "domains": domains,
        },
        "coverage_matrix": {"matrix": [], "coverage_pct": 0, "evidence_coverage_pct": 100},
        "report_quality_gate": {"status": "PASSED", "issues": []},
    }]
    book = load_workbook(BytesIO(structured_excel_report(
        "Тест", "15.2.4", docs, [], [], report_kind="manager", checklist_results=[],
    )), read_only=True, data_only=True)
    summary = {row[0].value: row[1].value for row in book["Резюме"].iter_rows()}
    assert summary["Статус отчёта"] == "Итоговый — проверка неполная"
    assert summary["Целостность отчёта"] == "Пройдена"
    assert summary["Готовность проверки"] == "Неполная"
    gate = list(book["Контроль отчёта"].iter_rows(values_only=True))
    assert gate[1][2] == "Пройдено"
    assert gate[2][2] == "Не завершено"
    book.close()
