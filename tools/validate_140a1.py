#!/usr/bin/env python3
"""End-to-end validation for ExpertCheck 14.0 Alpha 1."""

from __future__ import annotations

import argparse
import io
import json
import sys
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from analyzer import analyze_uploaded  # noqa: E402
from studio.data import structured_excel_report  # noqa: E402


VERSION = "14.0-alpha1-semantic-evidence-engine"
RAW_STATUS_CODES = {
    "ADMIT", "TRUSTED", "EXPERIMENTAL", "RETRIEVAL_ONLY", "PASSED", "FAILED",
    "BLOCKED", "NOT_REQUIRED", "SATISFIED", "UNSATISFIED", "VERIFIED_OK",
    "PROJECT_FINDING", "REVIEW_QUESTION", "SYSTEM_LIMITATION",
}


class Upload:
    def __init__(self, path: Path):
        self.path = path
        self.name = path.name
        self._buffer = io.BytesIO(path.read_bytes())

    def getvalue(self) -> bytes:
        return self.path.read_bytes()

    def read(self, *args, **kwargs):
        return self._buffer.read(*args, **kwargs)

    def seek(self, *args, **kwargs):
        return self._buffer.seek(*args, **kwargs)

    def tell(self):
        return self._buffer.tell()


def _counts(rows: list[dict], field: str) -> dict[str, int]:
    return dict(sorted(Counter(str(row.get(field) or "") for row in rows).items()))


def _localized_status_audit(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    leaks: list[dict[str, str]] = []
    book = load_workbook(path, read_only=True, data_only=True)
    for sheet in book.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, ())
        status_columns = [
            index for index, value in enumerate(header)
            if any(token in str(value or "").lower() for token in ("статус", "результат", "critic", "gate", "класс"))
        ]
        for row_index, row in enumerate(rows, 2):
            for column in status_columns:
                value = str(row[column] or "").strip() if column < len(row) else ""
                if value in RAW_STATUS_CODES:
                    leaks.append({"sheet": sheet.title, "cell": f"R{row_index}C{column+1}", "value": value})
    book.close()
    return leaks


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("--output", type=Path, default=ROOT / "VALIDATION_140_ALPHA1.json")
    parser.add_argument("--reports", type=Path, default=ROOT / "validation_reports_140a1")
    args = parser.parse_args()

    source_paths = sorted(path for path in args.source_dir.iterdir() if path.is_file() and path.suffix.lower() in {".pdf", ".xml"})
    documents, findings, comparisons = analyze_uploaded(
        [Upload(path) for path in source_paths], ROOT, ai_options={"enabled": False, "level": "off"}
    )
    first = documents[0] if documents else {}
    parents = list(first.get("assignment_compliance") or [])
    atoms = list(first.get("assignment_atomic_compliance") or [])
    checklist = [row for row in (first.get("automatic_checklist_review") or {}).get("results") or [] if not row.get("is_heading")]
    assignment_summary = dict(first.get("assignment_compliance_summary") or {})
    reconstruction = dict(first.get("evidence_reconstruction") or {})

    unsafe_checklist = [
        {
            "item": row.get("item_no") or row.get("position"),
            "typed_check": row.get("typed_check"),
            "status": row.get("status"),
        }
        for row in checklist
        if str(row.get("final_verification_kind") or row.get("verification_kind") or "").upper() == "VERIFIED_OK"
        and (
            str(row.get("typed_check") or "").upper() in {"SPECIALIST_REVIEW", "ENGINEERING_SEMANTIC_REVIEW", "NORMATIVE_CONTENT_REVIEW"}
            or row.get("automatic_verdict_eligible") is False
            or row.get("candidate_evidence_only")
        )
    ]
    compressor = [
        row for row in comparisons
        if "компресс" in str(row.get("object") or row.get("object_name") or "").lower()
        and str(row.get("parameter_code") or "").upper() == "AREA_BUILD"
        and (str(row.get("finding_type") or "").upper() == "PROJECT_FINDING" or "РАСХОЖД" in str(row.get("status") or "").upper())
    ]
    def numeric_value(row: dict) -> float | None:
        for key in ("value", "value_num", "normalized_value", "value_text"):
            raw = row.get(key)
            if raw in (None, ""):
                continue
            match = __import__('re').search(r"\d+(?:[,.]\d+)?", str(raw))
            if match:
                return float(match.group(0).replace(',', '.'))
        return None

    pump_height = [
        row for row in findings
        if str(row.get("parameter_code") or "").upper() == "HEIGHT_BUILD"
        and abs((numeric_value(row) or 0) - 25.0) < 0.001
        and "насосн" in str(row.get("object_hint") or row.get("object_name") or "").lower()
    ]

    args.reports.mkdir(parents=True, exist_ok=True)
    report_paths = []
    names = {
        "manager": "ExpertCheck_Резюме_руководителя_14.0A1.xlsx",
        "gip": "ExpertCheck_Отчёт_ГИПа_14.0A1.xlsx",
        "technical": "ExpertCheck_Техническое_приложение_14.0A1.xlsx",
    }
    for kind, name in names.items():
        path = args.reports / name
        path.write_bytes(structured_excel_report(
            "Контрольный комплект ДСК", VERSION, documents, findings, comparisons,
            report_kind=kind, checklist_results=checklist,
        ))
        report_paths.append(path)
    localization_leaks = [
        {"file": path.name, **leak}
        for path in report_paths for leak in _localized_status_audit(path)
    ]

    pipeline_errors = list(first.get("pipeline_errors") or [])
    summary = {
        "version": VERSION,
        "input": {"documents": len(documents), "files": [path.name for path in source_paths]},
        "output": {"findings": len(findings), "comparisons": len(comparisons), "pipeline_errors": pipeline_errors},
        "assignment": {
            "parent_requirements": len(parents), "atomic_conditions": len(atoms),
            "public_total": assignment_summary.get("total"),
            "public_coverage_pct": assignment_summary.get("automatic_coverage_pct"),
            "parent_verdicts": _counts(parents, "verification_kind"),
            "atomic_verdicts": _counts(atoms, "verification_kind"),
        },
        "checklist": {
            "rows": len(checklist), "statuses": _counts(checklist, "status"),
            "unsafe_automatic_closures": unsafe_checklist,
        },
        "evidence_reconstruction": reconstruction.get("summary") or {},
        "semantic_evidence_engine": first.get("semantic_evidence_engine") or {},
        "coverage_matrix": first.get("coverage_matrix") or {},
        "semantic_project_graph": (first.get("semantic_project_graph") or {}).get("summary") or {},
        "high_value_sanitization": first.get("high_value_sanitization_audit") or {},
        "known_regression_cases": {
            "compressor_area_mismatch_detected": bool(compressor),
            "pump_height_25_quarantined": bool(pump_height) and all(row.get("comparison_excluded") for row in pump_height),
        },
        "quality_gate": first.get("report_quality_gate") or {},
        "report_localization_leaks": localization_leaks,
        "reports": [path.name for path in report_paths],
    }
    args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    valid = (
        len(source_paths) == 12
        and not pipeline_errors
        and not unsafe_checklist
        and assignment_summary.get("total") == len(parents)
        and bool(compressor)
        and summary["known_regression_cases"]["pump_height_25_quarantined"]
        and not localization_leaks
        and (first.get("report_quality_gate") or {}).get("status") == "PASSED"
        and (first.get("coverage_matrix") or {}).get("evidence_ready", 0) >= (first.get("coverage_matrix") or {}).get("completed", 0)
    )
    return 0 if valid else 1


if __name__ == "__main__":
    raise SystemExit(main())
