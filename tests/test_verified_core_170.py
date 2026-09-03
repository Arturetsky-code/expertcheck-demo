from __future__ import annotations

import json
import io

from openpyxl import load_workbook

from core.ai_gateway import AIResult, GroqProvider, OpenRouterProvider
from core.provider_benchmark import BENCHMARK_VERSION, benchmark_cases, qualified_ranking, run_provider_benchmark
from core.verified_verdict_gate import enforce_verified_verdicts
from studio.data import structured_excel_report


def _decision(packet_id: str, verdict: str) -> dict:
    positive = verdict in {"SUPPORTS", "CONTRADICTS"}
    return {
        "packet_id": packet_id,
        "verdict": verdict,
        "evidence_ids": [f"{packet_id}-E1"] if positive else [],
        "same_entity": True if verdict != "OTHER_ENTITY" else False,
        "same_property": True if verdict != "OTHER_METRIC" else False,
        "qualifiers_satisfied": positive,
        "modality_satisfied": positive,
        "confidence": 0.99,
        "reason": "Синтетический квалификационный ответ.",
    }


class PerfectProvider:
    name = "Perfect"
    model = "fixed-model"
    api_key = "configured"

    def __init__(self, expected: dict[str, str]):
        self.expected = expected

    def generate_validated(self, prompt, system="", validator=None, json_schema=None):
        assert json_schema and json_schema["required"] == ["decisions"]
        payload = json.loads(prompt)
        decisions = [
            _decision(packet["packet_id"], self.expected[packet["packet_id"]])
            for packet in payload["packets"]
        ]
        text = json.dumps({"decisions": decisions}, ensure_ascii=False)
        assert validator is None or validator(text)
        return AIResult(
            True, self.name, text=text, status_code=200, model=self.model,
            latency_ms=25, schema_mode="STRICT_JSON_SCHEMA",
        )


def test_provider_benchmark_has_thirty_anonymous_cases_and_strict_gates():
    cases = benchmark_cases()
    assert len(cases) == 30
    assert all("ДСК" not in json.dumps(case, ensure_ascii=False) for case in cases)
    expected = {case["case_id"]: case["expected_verdict"] for case in cases}
    result = run_provider_benchmark(PerfectProvider(expected), repeats=3, batch_size=5)
    assert result["qualified"] is True
    assert result["metrics"]["request_success_pct"] == 100.0
    assert result["metrics"]["schema_adherence_pct"] == 100.0
    assert result["metrics"]["semantic_accuracy_pct"] == 100.0
    assert result["metrics"]["false_positive_pct"] == 0.0
    assert result["metrics"]["repeatability_pct"] == 100.0
    assert result["metrics"]["calls"] == 18


def test_provider_ranking_excludes_unqualified_candidate_and_selects_stable_winner():
    results = {
        "Groq": {
            "version": BENCHMARK_VERSION,
            "completed": True,
            "qualified": True,
            "metrics": {
                "request_success_pct": 100, "schema_adherence_pct": 100,
                "semantic_accuracy_pct": 98, "false_positive_pct": 0,
                "repeatability_pct": 100, "latency_p95_ms": 1200,
            },
        },
        "OpenRouter": {
            "version": BENCHMARK_VERSION,
            "completed": True,
            "qualified": False,
            "metrics": {
                "request_success_pct": 100, "schema_adherence_pct": 100,
                "semantic_accuracy_pct": 94, "false_positive_pct": 2,
                "repeatability_pct": 100, "latency_p95_ms": 800,
            },
        },
    }
    assert qualified_ranking(results) == ["Groq"]


def test_openrouter_free_is_rejected_from_verified_core():
    result = OpenRouterProvider("key", "openrouter/free").generate("test")
    assert result.ok is False
    assert result.status_code == 412
    assert "не допускается" in result.error


def test_groq_uses_strict_json_schema_without_unstructured_retry():
    provider = GroqProvider("key", "openai/gpt-oss-120b")
    provider._available_models_cache = ["openai/gpt-oss-120b"]
    payloads = []

    def fake_post(url, headers, payload, timeout=45):
        payloads.append(payload)
        return 200, {"choices": [{"message": {"content": '{"decisions":[]}'}}]}

    provider._post = fake_post
    result = provider.generate_structured(
        "{}", "Верните JSON", json_schema={
            "type": "object",
            "properties": {"decisions": {"type": "array", "items": {"type": "object"}}},
            "required": ["decisions"],
            "additionalProperties": False,
        },
    )
    assert result.ok is True
    assert result.schema_mode == "STRICT_JSON_SCHEMA"
    assert len(payloads) == 1
    schema = payloads[0]["response_format"]["json_schema"]
    assert schema["strict"] is True
    assert schema["schema"]["required"] == ["decisions"]


def test_final_gate_blocks_real_false_barrier_l5_even_if_previous_layers_passed():
    rows = [{
        "question": "Проверить наличие уширения обочин на участках устройства барьерного ограждения",
        "verification_kind": "VERIFIED_OK",
        "final_verification_kind": "VERIFIED_OK",
        "status": "Да",
        "evidence_level": "L5",
        "proof_kind": "VERIFIED_ENGINEERING_EVIDENCE",
        "checker_family": "Смысловые проектные решения",
        "checker_mode": "Независимый консенсус",
        "semantic_consensus_completed": 0,
        "adversarial_state": "PASSED",
        "verification_evidence": [{
            "document": "ПЗУ.pdf", "page": 27,
            "text": "Территория ограждается сетчатыми панелями.",
        }],
        "recommendation": "Дополнительное действие не требуется.",
    }]
    summary = enforce_verified_verdicts(rows, domain="checklist")
    assert summary["blocked"] == 1
    assert rows[0]["verification_kind"] == "REVIEW_QUESTION"
    assert rows[0]["evidence_level"] == "L4"
    assert rows[0]["verified_core_gate_state"] == "BLOCKED"
    assert "Смысловой маршрут" in " ".join(rows[0]["verified_core_gate_reasons"])


def test_final_gate_preserves_addressable_deterministic_numeric_verdict():
    rows = [{
        "requirement_text": "Продолжительность смены — 8 часов",
        "verification_kind": "VERIFIED_OK",
        "final_verification_kind": "VERIFIED_OK",
        "status": "Соответствует заданию",
        "evidence_level": "L5",
        "proof_kind": "STRUCTURED_VALUE",
        "adversarial_state": "PASSED",
        "verification_evidence": [{
            "document": "ТХ.pdf", "page": 10, "text": "Продолжительность смены 8 часов",
        }],
    }]
    summary = enforce_verified_verdicts(rows, domain="assignment")
    assert summary["passed"] == 1
    assert rows[0]["verification_kind"] == "VERIFIED_OK"
    assert rows[0]["evidence_level"] == "L5"
    assert rows[0]["verified_core_gate_state"] == "PASSED"


def test_false_barrier_l5_cannot_reappear_in_exported_gip_xlsx():
    false_row = {
        "question": "Проверить наличие уширения обочин на участках устройства барьерного ограждения",
        "automatic_section": "ПЗУ",
        "verification_kind": "VERIFIED_OK",
        "final_verification_kind": "VERIFIED_OK",
        "status": "Да",
        "evidence_level": "L5",
        "proof_kind": "VERIFIED_ENGINEERING_EVIDENCE",
        "checker_family": "Смысловые проектные решения",
        "checker_mode": "Независимый консенсус",
        "semantic_consensus_completed": 0,
        "adversarial_state": "PASSED",
        "verification_evidence": [{
            "document": "ПЗУ.pdf", "page": 27,
            "text": "Территория ограждается сетчатыми панелями.",
        }],
        "recommendation": "Дополнительное действие не требуется.",
    }
    checklist_review = {"results": [false_row]}
    document = {
        "Файл": "ПЗУ.pdf", "Тип документа": "ПЗУ", "Страниц": 1,
        "assignment_compliance": [], "normative_compliance_audit": [],
        "automatic_checklist_review": checklist_review,
        "project_review_plan": {}, "coverage_matrix": {},
        "semantic_evidence_engine": {}, "report_quality_gate": {"status": "PASSED"},
    }
    payload = structured_excel_report(
        "E2E Verified Core", "ExpertCheck 17.0 · Verified Core",
        [document], [], [], report_kind="gip", checklist_results=[false_row],
    )
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    rows = list(workbook["Чек-листы"].iter_rows(values_only=True))
    header = list(rows[0])
    exported = dict(zip(header, rows[1]))
    assert exported["Результат ExpertCheck"] == "Требует проверки"
    assert exported["Итоговый класс проверки"] == "Требует проверки специалистом"
    assert exported["Уровень доказательства"] != "L5 — строгая проверка завершена"
    assert exported["Код причины незавершения"] == "VERIFIED_CORE_FINAL_GATE_BLOCKED"
