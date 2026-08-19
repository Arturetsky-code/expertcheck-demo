
from pathlib import Path
import io
import pandas as pd
from openpyxl import load_workbook
from core.assignment_compliance import compare_requirements
from core.normative_requirement_analyzer import NormativeRequirementAnalyzer
from studio.data import structured_excel_report

ROOT=Path(__file__).parent

def test_semantic_similarity_never_auto_confirms():
    req=[{"requirement_id":"A1","source_document":"Задание.pdf","page":1,
          "requirement_text":"Обеспечить автоматический контроль уровня раствора и аварийную сигнализацию.",
          "object_name":"","parameter_code":"","required_value":None,"unit":"",
          "requirement_type":"SEMANTIC","confidence":.7}]
    findings=[{"document":"ТХ.pdf","page":10,"context":"Предусмотрен автоматический контроль уровня раствора в резервуаре.",
               "parameter_code":"","object_hint":"Резервуар"}]
    rows=compare_requirements(req,findings,[])
    assert rows[0]["status"]=="Требуется смысловая проверка"
    assert rows[0]["evidence_candidates"]
    assert "не доказывают" in rows[0]["decision_basis"]

def test_numeric_requires_same_object():
    req=[{"requirement_id":"A2","source_document":"Задание.pdf","page":1,
          "requirement_text":"Предусмотреть КПП площадью застройки 42,5 м2",
          "object_name":"Кпп","parameter_code":"AREA_BUILD","required_value":42.5,"unit":"м2",
          "requirement_type":"NUMERIC","confidence":.9}]
    findings=[{"document":"ПЗ.pdf","page":5,"parameter_code":"AREA_BUILD","parameter_name":"Площадь застройки",
               "value":42.5,"unit":"м²","object_hint":"Насосная","semantic_anchor_name":"Насосная"}]
    rows=compare_requirements(req,findings,[])
    assert rows[0]["status"]=="Требование не подтверждено"

def test_normative_analyzer_does_not_invent_clause_requirement():
    n=NormativeRequirementAnalyzer(ROOT/"knowledge")
    rows=n.analyze_page("ПЗ.pdf",5,"Расчет выполнен в соответствии с СП 999.99999.2020, п. 5.2.3. Требуемая величина принята по расчету.")
    assert rows
    assert rows[0]["clause"]=="5.2.3"
    assert rows[0]["curated_requirement"]==""
    assert "верификац" in rows[0]["analysis_status"].lower()

def test_report_checklists_are_grouped_by_section_and_headers_russian():
    docs=pd.DataFrame([{"Файл":"ПЗ.pdf","Тип документа":"ПЗ","normative_validity_audit":[],"assignment_compliance":[]}])
    checks=[
      {"automatic_section":"ИОС1","automatic_checklist":"Электрика","item_no":"1","question":"Проверить мощность","status":"Нет","evidence":"x"},
      {"automatic_section":"ПЗУ","automatic_checklist":"Генплан","item_no":"2","question":"Проверить проезд","status":"Требует проверки","evidence":"y"},
    ]
    payload=structured_excel_report("Тест","9.4",docs,pd.DataFrame([]),pd.DataFrame([]),report_kind="gip",risks=[],checklist_results=checks,assembly_rows_data=[])
    wb=load_workbook(io.BytesIO(payload),read_only=True)
    assert "Чек-листы — сводка" in wb.sheetnames
    assert any(x.startswith("ЧЛ ИОС1") for x in wb.sheetnames)
    assert any(x.startswith("ЧЛ ПЗУ") for x in wb.sheetnames)
    ws=wb["Чек-листы — сводка"]
    headers=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    assert "Раздел" in headers and "Чек-лист" in headers
    assert "automatic_section" not in headers
    wb.close()
