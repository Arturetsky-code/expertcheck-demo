from __future__ import annotations

from io import BytesIO
import json

from openpyxl import load_workbook

from core.ai_gateway import AIResult, FailoverProvider
from core.project_review_planner import build_review_plan
from core.project_snapshot import (
    SNAPSHOT_VERSION,
    build_analysis_snapshot,
    corpus_fingerprint,
    load_project_snapshot,
    project_snapshot_bytes,
    recheck_project_snapshot,
)
from core.report_quality_gate import validate_review_plan
from core.semantic_evidence_engine import _call_batches
from studio.data import reconcile_question_headline, structured_excel_report


class _Provider:
    name = "TestAI"
    api_key = "configured"

    def __init__(self):
        self.calls = 0

    def generate(self, prompt: str, system: str = "") -> AIResult:
        self.calls += 1
        packet_id = json.loads(prompt)["packets"][0]["packet_id"]
        return AIResult(True, self.name, text=json.dumps({
            "decisions": [{
                "packet_id": packet_id, "verdict": "INSUFFICIENT", "evidence_ids": [],
                "same_entity": False, "same_property": False,
                "qualifiers_satisfied": False, "modality_satisfied": False,
                "confidence": 0.2, "reason": "Доказательств недостаточно.",
            }],
        }), status_code=200, model="test-model")


class _InvalidVerdictProvider(_Provider):
    def generate(self, prompt: str, system: str = "") -> AIResult:
        self.calls += 1
        packet_id = json.loads(prompt)["packets"][0]["packet_id"]
        return AIResult(True, self.name, text=json.dumps({
            "decisions": [{
                "packet_id": packet_id, "verdict": "MAYBE", "evidence_ids": [],
                "same_entity": False, "same_property": False,
                "qualifiers_satisfied": False, "modality_satisfied": False,
                "confidence": 0.2, "reason": "Недопустимый тестовый вывод.",
            }],
        }), status_code=200, model="test-model")


def test_semantic_checkpoint_reuses_completed_packet_without_provider_call():
    checkpoint = {}
    provider = FailoverProvider([_Provider()])
    packets = [{"packet_id": "P-1522"}]
    first, errors, _calls = _call_batches(provider, packets, checkpoint=checkpoint)
    assert not errors and "P-1522" in first
    actual_provider = provider.providers[0]
    assert actual_provider.calls == 1
    second, errors, calls = _call_batches(provider, packets, checkpoint=checkpoint)
    assert not errors and second == first
    assert actual_provider.calls == 1
    assert calls[0]["state"] == "CHECKPOINT_REUSED"


def test_invalid_semantic_decision_is_not_persisted_in_checkpoint():
    checkpoint = {}
    provider = FailoverProvider([_InvalidVerdictProvider()])
    answers, _errors, _calls = _call_batches(
        provider, [{"packet_id": "P-BAD"}], checkpoint=checkpoint, retry_limit=1,
    )
    assert answers == {}
    assert checkpoint == {}


def test_report_headline_reconciles_to_exported_specialist_queue():
    rows = [
        ["Вопросов специалисту", 410],
        ["Причины неполноты", "вопросов специалисту: 366; вне автоматического покрытия: 253; AI-пакетов в очереди: 83"],
        ["Итоговый вывод", "Старый итог"],
    ]
    reconcile_question_headline(
        rows, exact_question_count=370, project_findings=1,
        system_limitations=253, fallback="—",
    )
    assert rows[0][1] == 370
    assert rows[1][1] == "вопросов специалисту: 370; вне автоматического покрытия: 253; AI-пакетов в очереди: 83"
    assert "370" in rows[2][1] and "253" in rows[2][1]


def test_xlsx_question_total_and_quality_gate_sheet_are_auditable():
    question = {
        "plan_id": "CHECK-1", "domain": "Чек-листы", "title": "Проверить схему",
        "verification_kind": "REVIEW_QUESTION", "status": "Требует проверки",
        "coverage_reason": "Нужна смысловая проверка", "evidence_level": "L4",
    }
    domains = {
        "Задание на проектирование": {"total": 0}, "НТД": {"total": 0},
        "Чек-листы": {"total": 1, "review": 1, "system_limitation": 0, "coverage_pct": 0},
    }
    docs = [{
        "Файл": "ИОС2.pdf", "Раздел": "ИОС2", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [question], "project_findings": 0, "review_questions": 99,
            "system_limitations": 0, "domains": domains,
        },
        "coverage_matrix": {"matrix": [], "coverage_pct": 0, "evidence_coverage_pct": 100},
        "report_quality_gate": {"status": "FAILED", "issues": ["Тестовая причина gate."]},
    }]
    book = load_workbook(BytesIO(structured_excel_report(
        "Тест", "15.2.2", docs, [], [], report_kind="gip", checklist_results=[],
    )), read_only=True, data_only=True)
    assert book["Вопросы специалисту"].max_row == 2
    summary = {row[0].value: row[1].value for row in book["Резюме"].iter_rows()}
    assert summary["Вопросов специалисту"] == 1
    gate_text = " ".join(str(cell.value or "") for row in book["Контроль отчёта"].iter_rows() for cell in row)
    assert "Тестовая причина gate." in gate_text
    book.close()


def test_quality_gate_accepts_addressable_values_rendered_in_sources():
    plan = build_review_plan(
        assignment_rows=[], normative_rows=[], checklist_review={"results": []},
    )
    gate = validate_review_plan(plan, comparisons=[{
        "status": "СОВПАДАЕТ", "parameter_code": "AREA_BUILD",
        "parameter_name": "Площадь застройки",
        "sources": "ПЗ, стр. 45: 89.9 м² | ПЗУ, стр. 18: 89.9 м²",
    }])
    assert gate["status"] == "PASSED"


def test_project_snapshot_roundtrip_contains_rerunnable_page_corpus():
    analysis = build_analysis_snapshot(
        [{"Файл": "ПЗ.pdf", "Тип документа": "ПЗ", "Страниц": 1}],
        page_corpus=[{"document": "ПЗ.pdf", "document_type": "ПЗ", "page": 1, "text": "Высота 2,5 м"}],
        fact_graph={"facts": [{"value": 2.5}], "summary": {"facts": 1}},
        object_registry=[{"Наименование объекта": "Насосная"}],
    )
    assert analysis["version"] == SNAPSHOT_VERSION
    assert analysis["summary"]["rerunnable_without_pdf"] is True
    documents = [{
        "Файл": "ПЗ.pdf", "core_version": "15.2.2",
        "analysis_snapshot": analysis,
        "assignment_atomic_compliance": [{"atom_id": "A-1"}],
    }]
    payload = load_project_snapshot(project_snapshot_bytes(documents, [], []))
    assert payload["analysis_snapshot"]["page_corpus"][0]["text"] == "Высота 2,5 м"
    assert payload["analysis_snapshot"]["quality_gate_inputs"]["object_registry"][0]["Наименование объекта"] == "Насосная"
    assert payload["assignment_atomic_compliance"][0]["atom_id"] == "A-1"


def test_project_snapshot_rechecks_quality_gate_without_source_pdf():
    plan = build_review_plan(
        assignment_rows=[], normative_rows=[], checklist_review={"results": []},
    )
    payload = {
        "core_version": "15.2.2-resumable-verification-snapshot",
        "analysis_snapshot": {"snapshot_id": "snapshot-1", "page_corpus": [{"text": "source"}]},
        "project_review_plan": plan,
        "automatic_checklist_review": {"results": []},
        "comparisons": [],
    }
    result = recheck_project_snapshot(payload)
    assert result["quality_gate"]["status"] == "PASSED"
    assert result["pages_reused"] == 1
    assert result["source_pdf_required"] is False


def test_corpus_fingerprint_changes_for_same_length_content():
    first = corpus_fingerprint([{"document": "A.pdf", "page": 1, "text": "25 m"}])
    second = corpus_fingerprint([{"document": "A.pdf", "page": 1, "text": "2,5m"}])
    assert first != second
