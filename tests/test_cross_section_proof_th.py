from pathlib import Path
import io

from openpyxl import load_workbook

from core.cross_section_consistency import build_cross_section_checks
from core.cross_section_verification import (
    qualify_cross_section_verdicts,
    technology_proof_summary,
)
from core.engineering_review_engine import CrossSectionDependencyEngine
from core.project_review_planner import build_review_plan
from core.verified_verdict_gate import enforce_verified_verdicts
from studio.data import structured_excel_report


ROOT = Path(__file__).resolve().parents[1]


def _fact(section, value, unit, *, code="DESIGN_CAPACITY", page=10, document=None):
    return {
        "document": document or f"{section}.pdf",
        "document_type": section,
        "section": section,
        "page": page,
        "parameter_code": code,
        "parameter_name": "Проектная мощность объекта",
        "value": value,
        "value_text": f"Проектная мощность {value} {unit}",
        "unit": unit,
        "object_hint": "Технологический комплекс",
        "semantic_anchor_name": "Технологический комплекс",
        "genplan_position": "1.1",
        "confidence": 0.98,
        "core2_confidence": 0.98,
        "binding_status": "POSITION_LOCKED",
        "evidence_quality_decision": "VERIFIED",
        "evidence_comparison_eligible": True,
        "evidence_mismatch_eligible": True,
        "evidence_trust_score": 95,
    }


def _qualified(*facts):
    rows = build_cross_section_checks(list(facts))
    CrossSectionDependencyEngine(ROOT / "knowledge").enrich_comparisons(rows)
    summary = qualify_cross_section_verdicts(rows)
    return rows, summary


def test_th_owner_and_pz_control_create_verified_l5_without_ai():
    rows, summary = _qualified(
        _fact("ТХ", 1.6, "млн.т/год"),
        _fact("ПЗ", 1600, "тыс.т/год", page=22),
    )
    assert len(rows) == 1
    row = rows[0]
    assert row["status"] == "СОВПАДАЕТ"
    assert row["final_verification_kind"] == "VERIFIED_OK"
    assert row["evidence_level"] == "L5"
    assert row["cross_section_gate_state"] == "PASSED"
    assert row["cross_section_gate"]["owner_present"] == ["ТХ"]
    assert row["cross_section_gate"]["control_present"] == ["ПЗ"]
    assert summary == {"version": "proof-th-cross-section-v1", "checked": 1, "passed": 1, "blocked": 0}
    assert enforce_verified_verdicts(rows, domain="comparison")["passed"] == 1


def test_th_owner_and_control_mismatch_is_addressable_project_finding():
    rows, _ = _qualified(
        _fact("ТХ", 1.6, "млн.т/год"),
        _fact("ПЗ", 1.4, "млн.т/год", page=22),
    )
    row = rows[0]
    assert row["status"] == "ПОТЕНЦИАЛЬНОЕ РАСХОЖДЕНИЕ"
    assert row["final_verification_kind"] == "PROJECT_FINDING"
    assert row["finding_type"] == "PROJECT_FINDING"
    assert len(row["verification_evidence"]) == 2
    assert {item["page"] for item in row["verification_evidence"]} == {10, 22}


def test_two_control_sections_cannot_replace_missing_th_owner():
    rows, summary = _qualified(
        _fact("ПЗ", 1600, "тыс.т/год"),
        _fact("ООС", 1600, "тыс.т/год", page=22),
    )
    row = rows[0]
    assert row["status"] == "СОВПАДАЕТ"
    assert row["final_verification_kind"] == "SYSTEM_LIMITATION"
    assert row["cross_section_gate_state"] == "BLOCKED"
    assert "раздел-владелец" in " ".join(row["cross_section_gate_reasons"])
    assert summary["passed"] == 0


def test_different_physical_units_never_form_false_comparison():
    rows, summary = _qualified(
        _fact("ТХ", 100, "т/ч"),
        _fact("ПЗ", 100, "м³/ч", page=22),
    )
    assert len(rows) == 2
    assert {row["unit"] for row in rows} == {"т/ч", "м³/ч"}
    assert all(row["status"] == "НЕДОСТАТОЧНО ДАННЫХ" for row in rows)
    assert summary["passed"] == 0


def test_missing_page_blocks_categorical_cross_section_result():
    rows, _ = _qualified(
        _fact("ТХ", 1600, "тыс.т/год", page=None),
        _fact("ПЗ", 1600, "тыс.т/год", page=22),
    )
    row = rows[0]
    assert row["final_verification_kind"] == "SYSTEM_LIMITATION"
    assert row["evidence_level"] != "L5"
    assert "двух адресных" in " ".join(row["cross_section_gate_reasons"])


def test_cross_section_is_first_class_review_domain_and_th_metric():
    rows, _ = _qualified(
        _fact("ТХ", 1600, "тыс.т/год"),
        _fact("ПЗ", 1600, "тыс.т/год", page=22),
    )
    enforce_verified_verdicts(rows, domain="comparison")
    plan = build_review_plan(
        assignment_rows=[], normative_rows=[], checklist_review={"results": []}, comparisons=rows,
    )
    assert plan["domains"]["comparison"]["total"] == 1
    assert plan["domains"]["comparison"]["completed"] == 1
    assert plan["domains"]["comparison"]["automatic_coverage_pct"] == 100.0
    assert plan["items"][0]["domain"] == "Межраздельная сверка"
    proof = technology_proof_summary(rows)
    assert proof["checks"] == 1
    assert proof["completed"] == 1
    assert proof["strict_coverage_pct"] == 100.0
    assert proof["addressable_evidence_pct"] == 100.0


def test_report_counts_cross_section_finding_once_and_exports_proof_route():
    rows, _ = _qualified(
        _fact("ТХ", 1.6, "млн.т/год"),
        _fact("ПЗ", 1.4, "млн.т/год", page=22),
    )
    enforce_verified_verdicts(rows, domain="comparison")
    plan = build_review_plan(
        assignment_rows=[], normative_rows=[], checklist_review={"results": []}, comparisons=rows,
    )
    document = {
        "Файл": "ТХ.pdf", "Тип документа": "ТХ", "Страниц": 1,
        "completeness_user_confirmed": True,
        "assignment_compliance": [], "normative_compliance_audit": [],
        "automatic_checklist_review": {"results": []},
        "project_review_plan": plan,
        "coverage_matrix": {}, "semantic_evidence_engine": {},
        "report_quality_gate": {"status": "PASSED", "issues": []},
    }
    gip = structured_excel_report(
        "Proof", "ExpertCheck Proof", [document], [], rows, report_kind="gip",
    )
    workbook = load_workbook(io.BytesIO(gip), data_only=True)
    summary = {row[0]: row[1] for row in workbook["Резюме"].iter_rows(values_only=True) if row[0]}
    assert summary["Подтверждённых несоответствий проекта"] == 1
    assert summary["Межраздельная сверка: завершено"] == 1
    assert summary["Межраздельная сверка: несоответствий"] == 1

    technical = structured_excel_report(
        "Proof", "ExpertCheck Proof", [document], [], rows, report_kind="technical",
    )
    workbook = load_workbook(io.BytesIO(technical), data_only=True)
    headers = [cell.value for cell in workbook["Тех_сверки"][1]]
    assert "owner_sections" in headers
    assert "control_sections" in headers
    assert "cross_section_gate_state" in headers
