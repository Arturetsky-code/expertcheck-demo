from __future__ import annotations

from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from core.normative_validity import NormativeValidityChecker
from core.review_queue import build_review_clusters
from core.typed_check_engine import execute_typed_check
from core.assignment_compliance import extract_requirements
from studio.data import structured_excel_report
from core.semantic_evidence_engine import run_semantic_evidence_engine
from tests.test_semantic_evidence_engine_140a1 import FakeProvider, _row


def test_review_questions_are_grouped_without_losing_detail():
    rows = [
        {
            "ID": "Q-1", "Контур": "Чек-листы", "Объект": "—",
            "Проверка": "Проверить схему 1", "Причина": "Нужна смысловая проверка",
            "Ожидаемые разделы": "ИОС2", "Уровень доказательства": "L3",
        },
        {
            "ID": "Q-2", "Контур": "Чек-листы", "Объект": "—",
            "Проверка": "Проверить схему 2", "Причина": "Нужна смысловая проверка",
            "Ожидаемые разделы": "ИОС2", "Уровень доказательства": "L4",
        },
        {
            "ID": "Q-3", "Контур": "Задание на проектирование", "Объект": "Компрессорная",
            "Проверка": "Проверить режим", "Причина": "Нужна смысловая проверка",
            "Ожидаемые разделы": "ТХ, ПЗ", "Уровень доказательства": "L2",
        },
    ]
    clusters = build_review_clusters(rows)
    assert len(clusters) == 2
    checklist = next(row for row in clusters if row["Контур"] == "Чек-листы")
    assert checklist["Количество вопросов"] == 2
    assert checklist["Приоритет"] == "Высокий"
    assert "Q-1" in checklist["ID вопросов"] and "Q-2" in checklist["ID вопросов"]


def test_failed_critic_preflight_falls_back_to_advisory_judge():
    class UnavailableCritic(FakeProvider):
        def test_connection(self):
            return SimpleNamespace(
                ok=False, provider=self.name, model="critic", status_code=403,
                error="Контрольная модель недоступна.",
            )

    row = _row()
    judge = FakeProvider("OpenRouter", "judge")
    critic = UnavailableCritic("Groq", "critic")
    audit = run_semantic_evidence_engine(
        [row], fact_graph={"facts": [], "passages": []}, level="extended", limit=10,
        judge_provider=judge, critic_provider=critic,
    )
    assert audit["execution_mode"] == "ADVISORY_JUDGE_ONLY"
    assert audit["judge_responses"] == audit["advisory_completed"] == 1
    assert audit["critic_responses"] == 0 and critic.last_payload is None
    assert row["semantic_consensus_state"] == "ADVISORY_ONLY"
    assert row["verification_kind"] == "REVIEW_QUESTION" and row["evidence_level"] == "L4"


def test_advisory_lane_is_bounded_and_reports_progress():
    rows = []
    for index in range(20):
        row = _row()
        row["atom_id"] = f"REQ-{index + 1}-A001"
        rows.append(row)
    events = []
    audit = run_semantic_evidence_engine(
        rows, fact_graph={"facts": [], "passages": []}, level="extended", limit=40,
        judge_provider=FakeProvider("Groq", "judge"), critic_provider=None,
        progress_callback=lambda role, completed, total: events.append((role, completed, total)),
    )
    assert audit["execution_mode"] == "ADVISORY_JUDGE_ONLY"
    assert audit["judge_candidates"] == 20
    assert audit["judge_selected"] == audit["advisory_limit"] == 12
    assert audit["not_selected"] == 8
    assert events[-1] == ("JUDGE", 12, 12)


def test_transport_failure_opens_semantic_circuit_breaker():
    class FailedProvider:
        name = "Groq"

        def __init__(self):
            self.calls = 0

        def generate(self, prompt, system=""):
            self.calls += 1
            return SimpleNamespace(
                ok=False, provider=self.name, model="test", status_code=0,
                error="connection reset by peer",
            )

    rows = []
    for index in range(20):
        row = _row()
        row["atom_id"] = f"REQ-{index + 1}-A001"
        rows.append(row)
    provider = FailedProvider()
    audit = run_semantic_evidence_engine(
        rows, fact_graph={"facts": [], "passages": []}, level="extended", limit=40,
        judge_provider=provider, critic_provider=None,
    )
    assert audit["judge_selected"] == 12
    assert provider.calls == 2
    assert len(audit["judge_calls"]) == 2
    assert audit["judge_calls"][-1]["state"] == "CIRCUIT_BREAKER"


def test_assignment_extractor_reuses_existing_page_corpus():
    class Upload:
        def __init__(self, name, document_type):
            self.name = name
            self.declared_document_type = document_type

        def getvalue(self):
            return b"not-a-real-pdf"

    files = [Upload("Пояснительная записка.pdf", "ПЗ"), Upload("Задание.pdf", "Задание")]
    corpus = [
        {"document": "Пояснительная записка.pdf", "page": 1, "text": "Текст пояснительной записки проекта."},
        {"document": "Задание.pdf", "page": 1, "text": "Задание на проектирование. Предусмотреть освещение территории металлическими опорами;"},
    ]

    def forbidden_reader(data, name):
        raise AssertionError(f"Повторное чтение PDF запрещено: {name}")

    rows = extract_requirements(files, forbidden_reader, page_corpus=corpus)
    assert rows
    assert all(row["source_document"] == "Задание.pdf" for row in rows)


def test_normative_registry_distinguishes_verified_unverified_and_missing():
    checker = NormativeValidityChecker("knowledge")
    verified = checker.check("Постановление Правительства РФ от 16.02.2008 № 87")
    recognised = checker.check("СП 20.13330.2016")
    missing = checker.check("ГОСТ 8267-93")
    assert verified["registry_match_state"] == "VERIFIED_STATUS"
    assert verified["coverage_status"] == "Проверено по реестру"
    assert recognised["registry_match_state"] == "MATCHED_UNVERIFIED"
    assert "статус не верифицирован" in recognised["coverage_status"]
    assert recognised["project_risk_applicable"] is False
    assert missing["registry_match_state"] == "NOT_IN_REGISTRY"
    assert missing["coverage_status"] == "Требует наполнения KB"


def test_addressable_structured_trace_closes_only_l1_presence():
    finding = {
        "document": "Раздел ПД №5_ИОС2.pdf", "page": 12,
        "context": "На листе приведена принципиальная схема системы водоснабжения.",
        "evidence_quality_decision": "VERIFIED", "binding_status": "ROW_LOCKED",
        "physical_trace_level": "ROW_TRACE",
    }
    l1 = execute_typed_check({
        "typed_check": "DOCUMENT_CONTENT_PRESENCE", "verification_level": "L1_PRESENCE",
        "evidence_terms": ["принципиальная", "схема", "водоснабжения"],
    }, [finding], [])
    assert l1 and l1["status"] == "Да" and l1["proof_kind"] == "STRUCTURED_PRESENCE"
    l4 = execute_typed_check({
        "typed_check": "DOCUMENT_CONTENT_PRESENCE", "verification_level": "L4_COMPLETENESS",
        "evidence_terms": ["принципиальная", "схема", "водоснабжения"],
    }, [finding], [])
    assert l4 and l4["status"] == "Не проверено системой"


def test_reports_expose_grouped_review_queue():
    question = {
        "plan_id": "CHECK-1", "domain": "Чек-листы", "title": "Проверить схему",
        "verification_kind": "REVIEW_QUESTION", "verification_state": "Требует проверки специалистом",
        "status": "Требует проверки", "coverage_reason": "Нужна смысловая проверка",
        "coverage_reason_code": "INDEPENDENT_SEMANTIC_CONFIRMATION_REQUIRED",
        "expected_sections": ["ИОС2"], "evidence_level": "L4",
    }
    domains = {
        "Задание на проектирование": {"total": 0}, "НТД": {"total": 0},
        "Чек-листы": {"total": 1, "review": 1, "system_limitation": 0, "coverage_pct": 0},
    }
    docs = [{
        "Файл": "ИОС2.pdf", "Раздел": "ИОС2", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [question], "project_findings": 0, "review_questions": 1,
            "system_limitations": 0, "domains": domains,
        },
        "coverage_matrix": {"matrix": [], "coverage_pct": 0, "evidence_coverage_pct": 100},
        "report_quality_gate": {"status": "PASSED"},
    }]
    payload = structured_excel_report("Тест", "15.1", docs, [], [], report_kind="gip", checklist_results=[])
    book = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert "Вопросы специалисту" in book.sheetnames
    assert "Очередь проверки" in book.sheetnames
    assert book["Очередь проверки"].max_row == 2
    book.close()
