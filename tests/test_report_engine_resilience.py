import io

from openpyxl import load_workbook

from core.report_engine import build_structured_report
from studio.data import structured_excel_report


def test_report_generation_accepts_non_string_finding_fields():
    comparison = {
        "comparison_id": "CMP-REPORT-NON-STRING",
        "finding_type": "REVIEW_QUESTION",
        "user_status": float("nan"),
        "object": 101,
        "parameter_name": float("nan"),
        "status": "Требует проверки",
        "applicability_proven": True,
    }

    report = build_structured_report("Проект", [{}], [comparison])

    assert len(report["problems"]) == 1
    assert report["problems"][0]["object"] == "101"
    assert report["problems"][0]["parameter"] == "Проверка"
    assert report["problems"][0]["status"] == "Требует проверки"


def test_report_generation_deduplicates_mixed_scalar_types_without_crashing():
    base = {
        "finding_type": "PROJECT_FINDING",
        "user_status": 422,
        "object": 7,
        "parameter_name": 12,
        "explicit_contradiction": True,
    }

    report = build_structured_report(
        "Проект",
        [{}],
        [dict(base, comparison_id="CMP-1"), dict(base, comparison_id="CMP-2")],
    )

    assert len(report["problems"]) == 1
    assert report["problems"][0]["object"] == "7"
    assert report["problems"][0]["parameter"] == "12"
    assert report["problems"][0]["status"] == "422"


def test_technical_report_accepts_mixed_cross_section_gate_reasons():
    document = {
        "Файл": "ТХ.pdf",
        "Тип документа": "ТХ",
        "Страниц": 1,
        "assignment_compliance": [],
        "normative_compliance_audit": [],
        "automatic_checklist_review": {"results": []},
        "project_review_plan": {},
        "coverage_matrix": {},
        "semantic_evidence_engine": {},
        "report_quality_gate": {"status": "PASSED", "issues": []},
    }
    comparison = {
        "comparison_id": "CMP-MIXED-GATE-REASONS",
        "object": "Технологический комплекс",
        "parameter_name": "Производительность",
        "status": "Требует проверки",
        "cross_section_gate_state": "BLOCKED",
        "cross_section_gate_reasons": [
            "Не найден контрольный раздел.",
            None,
            422,
            {"code": "MISSING_OWNER", "reason": "Не найден раздел-владелец."},
            ["вложенная", "диагностика"],
        ],
    }

    payload = structured_excel_report(
        "Проект", "ExpertCheck 17.1 Proof", [document], [], [comparison],
        report_kind="technical", checklist_results=[],
    )

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    rows = list(workbook["Тех_сверки"].iter_rows(values_only=True))
    exported = dict(zip(rows[0], rows[1]))
    reasons = exported["gate_reasons"]
    assert "Не найден контрольный раздел" in reasons
    assert "422" in reasons
    assert "MISSING_OWNER" in reasons
    assert "вложенная" in reasons
