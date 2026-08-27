#!/usr/bin/env python3
"""End-to-end validation for ExpertCheck 15.0 Alpha 1."""

from __future__ import annotations

import argparse
import io
import json
import sys
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from analyzer import analyze_uploaded  # noqa: E402
from studio.data import structured_excel_report  # noqa: E402


VERSION = "15.0-alpha1-executable-verification-engine"
RAW_STATUS_CODES = {
    "ADMIT", "TRUSTED", "EXPERIMENTAL", "RETRIEVAL_ONLY", "PASSED", "FAILED",
    "BLOCKED", "NOT_REQUIRED", "SATISFIED", "UNSATISFIED", "VERIFIED_OK",
    "PROJECT_FINDING", "REVIEW_QUESTION", "SYSTEM_LIMITATION",
    "ACCEPTED", "NOT_RUN", "PARTIAL", "CONSENSUS", "DETERMINISTIC",
    "SPECIALIST_REVIEW", "ENGINEERING_SEMANTIC_REVIEW", "FEATURE_PRESENCE",
    "NORMATIVE_CONTENT_REVIEW", "ATOMIC_PATTERN_PRESENCE",
}


class Upload:
    def __init__(self, path: Path):
        self.path = path
        self.name = path.name
        self._buffer = io.BytesIO(path.read_bytes())

    def getvalue(self) -> bytes:
        return self.path.read_bytes()

    def read(self, *args, **kwargs):
        return self._buffer.read(*args, **kwargs)

    def seek(self, *args, **kwargs):
        return self._buffer.seek(*args, **kwargs)

    def tell(self):
        return self._buffer.tell()


def _counts(rows: list[dict], field: str) -> dict[str, int]:
    return dict(sorted(Counter(str(row.get(field) or "") for row in rows).items()))


def _localized_status_audit(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    leaks: list[dict[str, str]] = []
    book = load_workbook(path, read_only=True, data_only=True)
    for sheet in book.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, ())
        status_columns = [
            index for index, value in enumerate(header)
            if any(token in str(value or "").lower() for token in ("статус", "результат", "critic", "gate", "класс"))
        ]
        for row_index, row in enumerate(rows, 2):
            for column in status_columns:
                value = str(row[column] or "").strip() if column < len(row) else ""
                if value in RAW_STATUS_CODES:
                    leaks.append({"sheet": sheet.title, "cell": f"R{row_index}C{column+1}", "value": value})
    book.close()
    return leaks


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("--output", type=Path, default=ROOT / "VALIDATION_150_ALPHA1.json")
    parser.add_argument("--reports", type=Path, default=ROOT / "validation_reports_150a1")
    parser.add_argument("--version", default=VERSION)
    parser.add_argument("--report-tag", default="15.0A1")
    args = parser.parse_args()
    run_version = str(args.version or VERSION)

    source_paths = sorted(path for path in args.source_dir.iterdir() if path.is_file() and path.suffix.lower() in {".pdf", ".xml"})
    documents, findings, comparisons = analyze_uploaded(
        [Upload(path) for path in source_paths], ROOT, ai_options={"enabled": False, "level": "off"}
    )
    first = documents[0] if documents else {}
    parents = list(first.get("assignment_compliance") or [])
    atoms = list(first.get("assignment_atomic_compliance") or [])
    checklist = [row for row in (first.get("automatic_checklist_review") or {}).get("results") or [] if not row.get("is_heading")]
    technology_checklist = [row for row in checklist if "Технологические решения" in str(row.get("source_file") or row.get("automatic_checklist") or "")]
    assignment_summary = dict(first.get("assignment_compliance_summary") or {})
    reconstruction = dict(first.get("evidence_reconstruction") or {})

    unsafe_checklist = [
        {
            "item": row.get("item_no") or row.get("position"),
            "typed_check": row.get("typed_check"),
            "status": row.get("status"),
        }
        for row in checklist
        if str(row.get("final_verification_kind") or row.get("verification_kind") or "").upper() == "VERIFIED_OK"
        and (
            str(row.get("typed_check") or "").upper() in {"SPECIALIST_REVIEW", "ENGINEERING_SEMANTIC_REVIEW", "NORMATIVE_CONTENT_REVIEW"}
            or row.get("automatic_verdict_eligible") is False
            or row.get("candidate_evidence_only")
        )
    ]
    compressor = [
        row for row in comparisons
        if "компресс" in str(row.get("object") or row.get("object_name") or "").lower()
        and str(row.get("parameter_code") or "").upper() == "AREA_BUILD"
        and (str(row.get("finding_type") or "").upper() == "PROJECT_FINDING" or "РАСХОЖД" in str(row.get("status") or "").upper())
    ]
    def numeric_value(row: dict) -> float | None:
        for key in ("value", "value_num", "normalized_value", "value_text"):
            raw = row.get(key)
            if raw in (None, ""):
                continue
            match = __import__('re').search(r"\d+(?:[,.]\d+)?", str(raw))
            if match:
                return float(match.group(0).replace(',', '.'))
        return None

    pump_height = [
        row for row in findings
        if str(row.get("parameter_code") or "").upper() == "HEIGHT_BUILD"
        and abs((numeric_value(row) or 0) - 25.0) < 0.001
        and "насосн" in str(row.get("object_hint") or row.get("object_name") or "").lower()
    ]

    l4_contract_violations = [
        row.get("atom_id") or row.get("requirement_id")
        for row in atoms + list((first.get("automatic_checklist_review") or {}).get("atomic_verification", {}).get("atoms") or [])
        if str(row.get("evidence_level") or "") == "L4"
        and (
            str(row.get("evidence_contract_state") or "") != "SATISFIED"
            or not list(row.get("verification_evidence") or [])
            or not list(row.get("evidence") or [])
        )
    ]

    # Exported reports must reflect explicit user decisions made after analysis.
    for document in documents:
        document["completeness_user_confirmed"] = True
        document["object_registry_confirmed"] = True

    args.reports.mkdir(parents=True, exist_ok=True)
    report_paths = []
    names = {
        "manager": f"ExpertCheck_Резюме_руководителя_{args.report_tag}.xlsx",
        "gip": f"ExpertCheck_Отчёт_ГИПа_{args.report_tag}.xlsx",
        "technical": f"ExpertCheck_Техническое_приложение_{args.report_tag}.xlsx",
    }
    for kind, name in names.items():
        path = args.reports / name
        path.write_bytes(structured_excel_report(
            "Контрольный комплект ДСК", run_version, documents, findings, comparisons,
            report_kind=kind, checklist_results=checklist,
        ))
        report_paths.append(path)
    localization_leaks = [
        {"file": path.name, **leak}
        for path in report_paths for leak in _localized_status_audit(path)
    ]
    from openpyxl import load_workbook
    workbook_sheets = {}
    report_confirmation = {}
    report_question_rows = {}
    report_atomic_headers = {}
    for path in report_paths:
        book = load_workbook(path, read_only=True, data_only=True)
        workbook_sheets[path.name] = list(book.sheetnames)
        if "Резюме" in book.sheetnames:
            rows = list(book["Резюме"].iter_rows(values_only=True))
            report_confirmation[path.name] = {
                str(row[0]): str(row[1]) for row in rows[1:] if len(row) >= 2 and row[0]
            }.get("Комплектность")
        report_question_rows[path.name] = (
            max(0, book["Вопросы специалисту"].max_row - 1)
            if "Вопросы специалисту" in book.sheetnames else 0
        )
        report_atomic_headers[path.name] = (
            [str(cell.value or "") for cell in next(book["Задание — атомарные условия"].iter_rows())]
            if "Задание — атомарные условия" in book.sheetnames else []
        )
        book.close()

    pipeline_errors = list(first.get("pipeline_errors") or [])
    summary = {
        "version": run_version,
        "input": {"documents": len(documents), "files": [path.name for path in source_paths]},
        "output": {"findings": len(findings), "comparisons": len(comparisons), "pipeline_errors": pipeline_errors},
        "assignment": {
            "parent_requirements": len(parents), "atomic_conditions": len(atoms),
            "public_total": assignment_summary.get("total"),
            "public_coverage_pct": assignment_summary.get("automatic_coverage_pct"),
            "parent_verdicts": _counts(parents, "verification_kind"),
            "atomic_verdicts": _counts(atoms, "verification_kind"),
        },
        "checklist": {
            "rows": len(checklist), "statuses": _counts(checklist, "status"),
            "technology_rows": len(technology_checklist),
            "unsafe_automatic_closures": unsafe_checklist,
        },
        "evidence_reconstruction": reconstruction.get("summary") or {},
        "semantic_evidence_engine": first.get("semantic_evidence_engine") or {},
        "l4_contract_violations": l4_contract_violations,
        "coverage_matrix": first.get("coverage_matrix") or {},
        "semantic_project_graph": (first.get("semantic_project_graph") or {}).get("summary") or {},
        "high_value_sanitization": first.get("high_value_sanitization_audit") or {},
        "known_regression_cases": {
            "compressor_area_mismatch_detected": bool(compressor),
            "pump_height_25_quarantined": bool(pump_height) and all(row.get("comparison_excluded") for row in pump_height),
        },
        "quality_gate": first.get("report_quality_gate") or {},
        "report_localization_leaks": localization_leaks,
        "report_sheets": workbook_sheets,
        "report_confirmation": report_confirmation,
        "report_question_rows": report_question_rows,
        "report_atomic_headers": report_atomic_headers,
        "reports": [path.name for path in report_paths],
    }
    args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    valid = (
        len(source_paths) == 12
        and not pipeline_errors
        and not unsafe_checklist
        # The curated pack contains 97 catalog rows; two are hierarchy headings.
        and len(technology_checklist) == 95
        and not l4_contract_violations
        and assignment_summary.get("total") == len(parents)
        and bool(compressor)
        and summary["known_regression_cases"]["pump_height_25_quarantined"]
        and not localization_leaks
        and all(value == "Подтверждена" for value in report_confirmation.values())
        and "AI — сводка" in workbook_sheets.get(names["gip"], [])
        and "AI — сводка" in workbook_sheets.get(names["technical"], [])
        and all(count > 0 for count in report_question_rows.values())
        and "Оператор условия" in report_atomic_headers.get(names["gip"], [])
        and "Значение проекта" in report_atomic_headers.get(names["technical"], [])
        and (first.get("report_quality_gate") or {}).get("status") == "PASSED"
        and (first.get("coverage_matrix") or {}).get("evidence_ready", 0) >= (first.get("coverage_matrix") or {}).get("completed", 0)
    )
    return 0 if valid else 1


if __name__ == "__main__":
    raise SystemExit(main())
