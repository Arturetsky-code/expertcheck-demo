from io import BytesIO

from openpyxl import load_workbook

from core.evidence_reconstruction import reconstruct_high_value_evidence, sanitize_high_value_facts
from core.ru_labels import ru_label
from studio.data import structured_excel_report


def test_density_values_are_reconstructed_as_material_scoped_question():
    pages = [
        {"document": "ПЗ.pdf", "document_type": "ПЗ", "page": 7, "text": "Для руды насыпная плотность материала составляет 1,54 т/м3."},
        {"document": "ТХ.pdf", "document_type": "ТХ", "page": 32, "text": "Насыпная плотность материала — 1,5 т/м³ для руды."},
    ]
    payload = reconstruct_high_value_evidence(pages)
    assert payload["summary"]["density_facts"] == 2
    assert len(payload["comparisons"]) == 1
    check = payload["comparisons"][0]
    assert check["finding_type"] == "REVIEW_QUESTION"
    assert "ПЗ.pdf, стр. 7" in check["sources"]
    assert all(row["evidence_scope"] == "MATERIAL_OR_PROCESS" for row in payload["facts"])


def test_unitless_surface_number_cannot_become_flow_rate():
    rows = [{"parameter_code": "FLOW_RATE", "value": 49841, "unit": "", "fact_admission_decision": "ADMIT"}]
    audit = sanitize_high_value_facts(rows)
    assert audit["flow_false_positives_blocked"] == 1
    assert rows[0]["fact_admission_decision"] == "REJECT"
    assert rows[0]["comparison_excluded"] is True


def test_multi_value_voltage_and_table_order_moisture_are_reconstructed():
    pages = [{
        "document": "ТХ.pdf", "document_type": "ТХ", "page": 12,
        "text": "Влажность руды, % 6,8. Напряжение 6/0,4 кВ. Установленная мощность 20 кВт. Рабочее давление 8 бар. Аккумулирующая емкость №1 Объем 150 м3.",
    }]
    facts = reconstruct_high_value_evidence(pages)["facts"]
    by_code = {row["parameter_code"]: row for row in facts}
    assert by_code["MOISTURE"]["value"] == 6.8
    assert by_code["VOLTAGE"]["normalized_values"] == [6.0, 0.4]
    assert {"POWER_INSTALLED", "PRESSURE", "RES_VOLUME"} <= set(by_code)


def test_core_status_codes_have_russian_labels():
    for code in ("ADMIT", "TRUSTED", "RETRIEVAL_ONLY", "PASSED", "BLOCKED", "SATISFIED", "TEXT_OR_TABLE"):
        assert ru_label(code) != code


def test_report_status_cells_are_localized():
    docs = [{
        "Файл": "ПЗ.pdf", "Раздел": "ПЗ", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [], "project_findings": 0, "review_questions": 0, "system_limitations": 0,
            "domains": {
                "Задание на проектирование": {"total": 0}, "НТД": {"total": 0}, "Чек-листы": {"total": 0},
                "assignment": {"total": 0}, "normative": {"total": 0}, "checklist": {"total": 0},
            },
        },
        "report_quality_gate": {"status": "PASSED"},
        "assignment_atomic_compliance": [{
            "atom_id": "A1", "parent_requirement_id": "R1", "atom_text": "Проверка",
            "recipe_status": "RETRIEVAL_ONLY", "critic_state": "PASSED", "regression_state": "BLOCKED",
            "evidence_contract_state": "UNSATISFIED", "status": "Не проверено системой",
        }],
    }]
    book = load_workbook(BytesIO(structured_excel_report("Тест", "13.0", docs, [], [], report_kind="technical", checklist_results=[])), read_only=True, data_only=True)
    values = " ".join(str(cell.value or "") for sheet in book.worksheets for row in sheet.iter_rows() for cell in row)
    book.close()
    for raw in ("RETRIEVAL_ONLY", "PASSED", "BLOCKED", "UNSATISFIED"):
        assert raw not in values
