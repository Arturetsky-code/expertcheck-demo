#!/usr/bin/env python3
"""Reproducible end-to-end validation for ExpertCheck 11.1 Alpha 1."""

from __future__ import annotations

import argparse
import io
import json
import sys
from collections import Counter
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from analyzer import analyze_uploaded  # noqa: E402
from studio.data import structured_excel_report  # noqa: E402


VERSION = "11.1.0-alpha1-evidence-contract-engine"


class Upload:
    def __init__(self, path: Path):
        self.path = path
        self.name = path.name
        self._buffer = io.BytesIO(path.read_bytes())

    def getvalue(self) -> bytes:
        return self.path.read_bytes()

    def read(self, *args, **kwargs):
        return self._buffer.read(*args, **kwargs)

    def seek(self, *args, **kwargs):
        return self._buffer.seek(*args, **kwargs)

    def tell(self):
        return self._buffer.tell()


def _counts(rows: list[dict], field: str = "status") -> dict[str, int]:
    return dict(sorted(Counter(str(row.get(field) or "") for row in rows).items()))


def _locators(row: dict) -> list[str]:
    result = []
    for item in row.get("verification_evidence") or []:
        document = str(item.get("document") or "")
        page = item.get("page")
        section = str(item.get("section") or item.get("document_type") or "")
        if document and page not in (None, "") and section:
            result.append(f"{document}, стр. {page} [{section}]")
    return result


def _report_validation(report_dir: Path, finding_ids: list[str]) -> dict:
    from openpyxl import load_workbook

    result = {"files": {}, "all_findings_in_issues": True, "all_findings_in_actions": True}
    for path in sorted(report_dir.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets = list(workbook.sheetnames)
        all_values: dict[str, str] = {}
        for sheet in sheets:
            values = []
            for row in workbook[sheet].iter_rows(values_only=True):
                values.extend(str(value or "") for value in row)
            all_values[sheet] = "\n".join(values)
        issue_blob = all_values.get("Несоответствия и вопросы", "")
        action_blob = all_values.get("План действий", "") + all_values.get("Приоритетные действия", "")
        missing_issues = [item for item in finding_ids if item not in issue_blob]
        missing_actions = [item for item in finding_ids if item not in action_blob]
        result["files"][path.name] = {
            "sheets": sheets,
            "atomic_sheet": "Задание — атомарные условия" in sheets,
            "missing_findings_in_issues": missing_issues,
            "missing_findings_in_actions": missing_actions,
        }
        if path.name.find("Резюме") < 0:
            result["all_findings_in_issues"] &= not missing_issues
            result["all_findings_in_actions"] &= not missing_actions
        workbook.close()
    return result


def build_summary(documents: list[dict], findings: list[dict], comparisons: list[dict]) -> dict:
    first = documents[0] if documents else {}
    parents = list(first.get("assignment_compliance") or [])
    atoms = list(first.get("assignment_atomic_compliance") or [])
    assignment_summary = dict(first.get("assignment_compliance_summary") or {})
    checklist = [row for row in list((first.get("automatic_checklist_review") or {}).get("results") or []) if not row.get("is_heading")]
    checklist_atomic = dict((first.get("automatic_checklist_review") or {}).get("atomic_verification") or {})
    categorical = [row for row in atoms if row.get("verification_kind") in {"VERIFIED_OK", "PROJECT_FINDING"}]
    atom_findings = [row for row in atoms if row.get("verification_kind") == "PROJECT_FINDING"]
    bad_categorical = [
        row.get("atom_id") for row in categorical
        if not _locators(row)
        or row.get("critic_state") != "PASSED"
        or row.get("regression_state") != "PASSED"
        or row.get("semantic_gate_state") != "PASSED"
    ]
    pipeline_errors = []
    for document in documents:
        for error in document.get("pipeline_errors") or []:
            if error not in pipeline_errors:
                pipeline_errors.append(error)
    return {
        "version": VERSION,
        "input": {"documents": len(documents), "document_types": _counts(documents, "Тип документа")},
        "output": {"findings": len(findings), "comparisons": len(comparisons), "pipeline_errors": pipeline_errors},
        "assignment": {
            "source_rows": len(parents), "atomic_conditions": len(atoms),
            "additional_conditions": max(0, len(atoms) - len(parents)),
            "parent_statuses": _counts(parents), "atomic_statuses": _counts(atoms),
            "atomic_verdicts": _counts(atoms, "verification_kind"),
            "atomic_kinds": _counts(atoms, "atomic_kind"),
            "summary": assignment_summary,
            "categorical_without_complete_gate": bad_categorical,
            "findings": [{
                "atom_id": row.get("atom_id"), "parent_id": row.get("parent_requirement_id"),
                "condition": row.get("atom_text"), "difference": row.get("difference"),
                "evidence": _locators(row), "critic": row.get("critic_state"),
                "regression": row.get("regression_state"), "deep_evidence": row.get("deep_evidence_state"),
            } for row in atom_findings],
        },
        "checklists": {
            "rows": len(checklist), "statuses": _counts(checklist),
            "verdicts": _counts(checklist, "verification_kind"),
            "atomic": checklist_atomic.get("summary") or {},
        },
        "cross_section": {"total": len(comparisons), "statuses": _counts(comparisons)},
        "quality_gate": dict(first.get("report_quality_gate") or {}),
        "deep_evidence": dict(first.get("deep_evidence_review") or {}).get("metrics") or {},
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", type=Path)
    parser.add_argument("--output", type=Path, default=ROOT / "VALIDATION_111_ALPHA1.json")
    parser.add_argument("--reports", type=Path)
    args = parser.parse_args()
    source_paths = sorted(path for path in args.source_dir.iterdir() if path.is_file() and path.suffix.lower() in {".pdf", ".xml"})
    if not source_paths:
        raise SystemExit(f"No PDF/XML files found in {args.source_dir}")
    documents, findings, comparisons = analyze_uploaded([Upload(path) for path in source_paths], ROOT, ai_options={"enabled": False})
    summary = build_summary(documents, findings, comparisons)
    if args.reports:
        args.reports.mkdir(parents=True, exist_ok=True)
        checklist = list((documents[0].get("automatic_checklist_review") or {}).get("results") or []) if documents else []
        names = {
            "manager": "ExpertCheck_Резюме_руководителя_11.1A1.xlsx",
            "gip": "ExpertCheck_Отчёт_ГИПа_11.1A1.xlsx",
            "technical": "ExpertCheck_Техническое_приложение_11.1A1.xlsx",
        }
        for kind, name in names.items():
            payload = structured_excel_report(
                "Контрольный комплект ДСК", VERSION, documents, findings, comparisons,
                report_kind=kind, checklist_results=checklist,
            )
            (args.reports / name).write_bytes(payload)
        finding_ids = [str(row.get("atom_id")) for row in (documents[0].get("assignment_atomic_compliance") or []) if row.get("verification_kind") == "PROJECT_FINDING"]
        summary["reports"] = _report_validation(args.reports, finding_ids)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    report_ok = not args.reports or (summary.get("reports", {}).get("all_findings_in_issues") and summary.get("reports", {}).get("all_findings_in_actions"))
    valid = (
        summary["quality_gate"].get("status") == "PASSED"
        and not summary["output"]["pipeline_errors"]
        and not summary["assignment"]["categorical_without_complete_gate"]
        and report_ok
    )
    return 0 if valid else 1


if __name__ == "__main__":
    raise SystemExit(main())
