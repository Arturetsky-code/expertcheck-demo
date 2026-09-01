from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from core.atomic_verification_engine import verify_checklist_rows
from core.coverage_acceleration import coverage_budget
from core.deep_evidence_intelligence import compact_deep_evidence_review, run_deep_evidence_review
from core.semantic_evidence_engine import run_semantic_evidence_engine
from tests.test_semantic_evidence_engine_140a1 import FakeProvider, _row
from studio.data import structured_excel_report


def _semantic_rows(count: int):
    rows = []
    for index in range(count):
        row = _row()
        row["atom_id"] = f"CHECK-RUNTIME-{index + 1:03d}"
        rows.append(row)
    return rows


def test_initial_checklist_ai_is_deferred_but_full_target_is_preserved():
    budget = coverage_budget("extended", "extended")
    assert budget.checklist_semantic_limit >= 544
    assert budget.initial_checklist_semantic_limit == 0
    assert 0 < budget.continuation_checklist_batch_limit <= 20

    rows = _semantic_rows(5)
    judge = FakeProvider("OpenRouter", "judge")
    critic = FakeProvider("Groq", "critic")
    audit = run_semantic_evidence_engine(
        rows, fact_graph={"facts": [], "passages": []},
        level="off", limit=0, judge_provider=judge, critic_provider=critic,
    )
    assert audit["judge_candidates"] == 5
    assert audit["judge_selected"] == 0
    assert audit["not_selected"] == 5
    assert judge.last_payload is None and critic.last_payload is None


def test_resumable_batch_selects_next_packets_after_checkpoint():
    checkpoint = {}
    first_rows = _semantic_rows(5)
    first = run_semantic_evidence_engine(
        first_rows, fact_graph={"facts": [], "passages": []},
        level="extended", limit=2,
        judge_provider=FakeProvider("OpenRouter", "judge"),
        critic_provider=FakeProvider("Groq", "critic"),
        checkpoint=checkpoint,
    )
    assert first["judge_selected"] == first["judge_responses"] == 2
    assert first["judge_attempted"] == 2
    assert first["not_selected"] == 3

    second_rows = _semantic_rows(5)
    second = run_semantic_evidence_engine(
        second_rows, fact_graph={"facts": [], "passages": []},
        level="extended", limit=2,
        judge_provider=FakeProvider("OpenRouter", "judge"),
        critic_provider=FakeProvider("Groq", "critic"),
        checkpoint=checkpoint,
    )
    assert second["judge_selected"] == second["judge_responses"] == 4
    assert second["judge_checkpoint_reused"] == 2
    assert second["judge_attempted"] == 2
    assert second["not_selected"] == 1


def test_checklist_parent_does_not_duplicate_full_atomic_conditions():
    checklist = [{
        "question": "Проверить наличие значения производительности",
        "automatic_section": "ТХ",
        "compiled_rule": {
            "typed_check": "ENGINEERING_PARAMETER_PRESENCE",
            "verification_level": "L2_VALUE",
            "parameter_codes": ["CAPACITY"],
        },
        "typed_check": "ENGINEERING_PARAMETER_PRESENCE",
    }]
    fact = {
        "property_code": "CAPACITY", "value": 1600, "unit": "тыс. т/год",
        "document": "ТХ.pdf", "page": 17, "document_type": "ТХ",
        "source_trace": "Производительность комплекса 1600 тыс. т/год",
        "fact_admission_decision": "ADMIT", "binding_status": "ROW_LOCKED",
    }
    result = verify_checklist_rows(
        checklist, knowledge_root="knowledge",
        fact_graph={"facts": [fact], "passages": []}, page_corpus=[],
    )
    assert result["atoms"]
    assert "atomic_conditions" not in checklist[0]
    assert checklist[0]["atomic_condition_ids"]


def test_persisted_deep_review_does_not_duplicate_page_corpus():
    review = run_deep_evidence_review(
        [{"plan_id": "CHECK-1", "title": "Проверить производительность"}],
        page_corpus=[{
            "document": "ТХ.pdf", "document_type": "ТХ", "page": 17,
            "text": "Производительность комплекса 1600 тыс. т/год " * 100,
        }],
    )
    assert review["evidence_db"]["records"]
    record_count = review["evidence_db"]["record_count"]
    compact = compact_deep_evidence_review(review)
    assert "evidence_db" not in compact
    assert compact["evidence_db_summary"]["record_count"] == record_count
    assert all("evidence_candidates" not in row for row in compact["results"])


def test_deferred_checklist_queue_keeps_report_incomplete():
    domains = {
        "Задание на проектирование": {"total": 0},
        "НТД": {"total": 0},
        "Чек-листы": {"total": 0},
    }
    docs = [{
        "Файл": "ТХ.pdf", "Раздел": "ТХ", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [], "project_findings": 0, "review_questions": 0,
            "system_limitations": 0, "domains": domains,
        },
        "coverage_matrix": {"matrix": [], "coverage_pct": 0, "evidence_coverage_pct": 0},
        "report_quality_gate": {"status": "PASSED", "issues": []},
        "semantic_evidence_engine": {
            "assignment": {},
            "checklist": {"judge_candidates": 3, "judge_selected": 0, "not_selected": 3},
        },
    }]
    workbook = load_workbook(BytesIO(structured_excel_report(
        "Тест", "15.2.5", docs, [], [], report_kind="manager", checklist_results=[],
    )), read_only=True, data_only=True)
    summary = {row[0].value: row[1].value for row in workbook["Резюме"].iter_rows()}
    assert summary["Статус отчёта"] == "Итоговый — проверка неполная"
    assert summary["Готовность проверки"] == "Неполная"
    assert "AI-пакетов в очереди: 3" in summary["Причины неполноты"]
    workbook.close()
