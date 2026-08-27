from __future__ import annotations
import io
import json
import math
import re
from datetime import datetime, date
import pandas as pd
from core.ru_labels import ru_label, ru_join
from core.display_localization import parameter_label, status_label, scope_label, header_label, localize_service_value
from core.report_engine import build_decision_report, build_structured_report
from core.evidence_registry import build_evidence_index
from core.object_intelligence import build_object_decisions
from core.project_review_planner import build_review_plan
from core.verification_core import verification_label
from core.global_finding_gate import classify_finding
from core.project_assembly import (
    build_assembly_rows, filter_comparisons_by_keys, filter_passports_by_keys,
    filter_registry_by_keys, selected_keys,
)


def frames(result):
    if not result: return pd.DataFrame(),pd.DataFrame(),pd.DataFrame()
    d,f,c=result; return pd.DataFrame(d),pd.DataFrame(f),pd.DataFrame(c)

def status_group(value: str) -> str:
    t=str(value or '').upper()
    if 'РАСХОЖД' in t or 'КОНФЛИКТ' in t:return 'bad'
    if any(x in t for x in ('УТОЧ','НЕДОСТАТОЧ','НЕ ПРОВЕРЕНО','ПРЕДВАРИТ','ЧАСТИЧ','КАНДИДАТ')):return 'warn'
    if 'СОВПАД' in t or 'ПОДТВЕРЖ' in t:return 'ok'
    return 'info'


def _evidence_sources(row) -> str:
    """Build a visible document/page trace from all supported evidence shapes."""
    direct=row.get('sources') or row.get('Источники')
    if isinstance(direct,str) and direct.strip():
        return direct.strip()
    refs=[]
    if isinstance(direct,list):
        refs.extend(direct)
    for key in ('deep_evidence_candidates','evidence_candidates','evidence_records','supporting_evidence','matched_findings'):
        value=row.get(key) or []
        if isinstance(value,list): refs.extend(value)
    document=row.get('source_document') or row.get('document')
    if document:
        refs.append({'document':document,'page':row.get('source_page') or row.get('page')})
    rendered=[];seen=set()
    for ref in refs:
        if isinstance(ref,str): text=ref.strip()
        elif isinstance(ref,dict):
            doc=ref.get('document') or ref.get('source_document') or ref.get('source') or ''
            page=ref.get('page') or ref.get('source_page') or ''
            text=f"{doc}, стр. {page}" if doc and page else str(doc or '')
        else: text=''
        if text and text not in seen:
            seen.add(text); rendered.append(text)
    return ' | '.join(rendered) if rendered else 'Источник не сформирован — требуется целевой поиск'


def _ai_error_summary(errors) -> str:
    """Render compact error categories; raw provider text stays in AI calls."""
    categories={}
    for error in errors or []:
        low=str(error or '').lower()
        if '403' in low or 'forbidden' in low or 'permission' in low or 'запрещ' in low:
            label='Доступ к модели'
        elif '429' in low or 'rate limit' in low or 'лимит' in low:
            label='Лимит запросов'
        elif '413' in low or 'too large' in low or 'слишком больш' in low:
            label='Размер пакета'
        elif 'json' in low:
            label='Формат JSON'
        elif 'timeout' in low or 'timed out' in low or 'время ожидания' in low:
            label='Тайм-аут'
        elif '401' in low or 'api-ключ' in low or 'api key' in low:
            label='API-ключ'
        else:
            label='Прочая ошибка провайдера'
        categories[label]=categories.get(label,0)+1
    return '; '.join(f'{label}: {count}' for label,count in categories.items()) or 'Нет'


def _difference_field(row, key, default=''):
    value=row.get('difference') or {}
    return value.get(key,default) if isinstance(value,dict) else default

def metrics(df):
    if df.empty or 'status' not in df:return {'total':0,'ok':0,'warn':0,'bad':0}
    g=df['status'].map(status_group);return {'total':len(df),'ok':int(g.eq('ok').sum()),'warn':int(g.eq('warn').sum()),'bad':int(g.eq('bad').sum())}

def engineer_findings(df):
    if not isinstance(df,pd.DataFrame): df=pd.DataFrame(df or [])
    if df.empty:return df.copy()
    excluded={'PROJECT_NAME','PROJECT_CODE','PROJECT_YEAR','ISSUE_AUTHOR','CHIEF_ENGINEER','SIGNER','DOCUMENT_CODE','DOCUMENT_YEAR','XML_SCHEMA','FILE_NAME','FILE_CHECKSUM','OBJECT_ENTRY','OBJECT_CANDIDATE'}
    out=df.copy()
    if 'parameter_code' in out:out=out[~out['parameter_code'].fillna('').astype(str).isin(excluded)]
    return out

def _first_document_record(docs):
    if hasattr(docs,'empty'):
        return {} if docs.empty else docs.iloc[0].to_dict()
    if isinstance(docs,list) and docs:
        return dict(docs[0] or {})
    return {}

def composition_baseline(docs):
    return pd.DataFrame(_first_document_record(docs).get('composition_baseline') or [])

def raw_registry(docs):
    first=_first_document_record(docs)
    if not first:return pd.DataFrame()
    # Structured project composition is the fail-safe engineer-facing registry.
    # If present, it cannot be replaced by generic narrative extraction results.
    base = composition_baseline(docs)
    if not base.empty:
        consolidated = pd.DataFrame(first.get('consolidated_registry') or [])
        if not consolidated.empty and 'Позиция по ГП' in consolidated.columns:
            extra_cols=[c for c in consolidated.columns if c not in base.columns]
            if extra_cols:
                enrich=consolidated[['Позиция по ГП']+extra_cols].drop_duplicates('Позиция по ГП')
                base=base.merge(enrich,on='Позиция по ГП',how='left')
        return base
    return pd.DataFrame(first.get('consolidated_registry') or [])

def raw_candidates(docs):
    return pd.DataFrame(_first_document_record(docs).get('consolidated_candidates') or [])

def raw_passports(docs):
    return _first_document_record(docs).get('object_passports') or []

def assembly_rows(docs, findings=None):
    finding_rows=(findings.to_dict('records') if hasattr(findings,'to_dict') else findings) or []
    evidence_index=build_evidence_index(finding_rows)
    intelligence=build_object_decisions(finding_rows)
    trusted=raw_registry(docs).to_dict('records')
    # With an explicit composition baseline, generic text candidates belong only
    # to developer diagnostics. The primary table must show the actual project
    # composition, not sentences, TEP labels or section headings.
    candidates=[] if not composition_baseline(docs).empty else raw_candidates(docs).to_dict('records')
    return build_assembly_rows(trusted, candidates, evidence_index, intelligence)

def apply_project_assembly(docs, passports, comparisons, state_rows, confirmed):
    if not confirmed:
        return pd.DataFrame(), [], pd.DataFrame()
    allowed=selected_keys(state_rows)
    reg=pd.DataFrame(filter_registry_by_keys(raw_registry(docs).to_dict('records'),allowed))
    pas=filter_passports_by_keys(passports,allowed)
    cmp=pd.DataFrame(filter_comparisons_by_keys(comparisons.to_dict('records'),state_rows,allowed))
    return reg,pas,cmp

def registry(docs): return raw_registry(docs)
def passports(docs): return raw_passports(docs)

_ILLEGAL_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_EXCEL_MAX_TEXT = 32000


def _excel_safe_value(value):
    """Convert arbitrary diagnostic values to valid, compact XLSX cell values."""
    if value is None:
        return ''
    if isinstance(value, float):
        return '' if math.isnan(value) or math.isinf(value) else value
    if isinstance(value, (int, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, (dict, list, tuple, set)):
        try:
            value = json.dumps(value, ensure_ascii=False, default=str, separators=(',', ': '))
        except Exception:
            value = str(value)
    text = _ILLEGAL_XML_CHARS.sub('', str(value)).replace('\x00', '').strip()
    # Do not allow project text to be interpreted as an Excel formula.
    if text.startswith(('=', '+', '-', '@')):
        text = "'" + text
    return text[:_EXCEL_MAX_TEXT]


def _excel_safe_frame(frame: pd.DataFrame, *, columns: list[str] | None = None, max_rows: int | None = None) -> pd.DataFrame:
    if frame is not None and not isinstance(frame,pd.DataFrame):
        frame=pd.DataFrame(frame)
    if frame is None or frame.empty:
        return pd.DataFrame(columns=columns or [])
    result = frame.copy()
    if columns is not None:
        available = [column for column in columns if column in result.columns]
        result = result[available]
    if max_rows is not None:
        result = result.head(max_rows)
    result.columns = [_excel_safe_value(column) for column in result.columns]
    for column in result.columns:
        result[column] = result[column].map(_excel_safe_value)
    return result


def _compact_technical_frames(docs, findings, comparisons, report):
    comparison_items=comparisons.to_dict('records') if hasattr(comparisons,'to_dict') else list(comparisons or [])
    registry_columns = [
        'Позиция по ГП','Наименование объекта','Статус проектирования','Статус',
        'Источники','Страницы','Уверенность','Решение инспектора','Причины решения',
    ]
    comparison_columns = [
        'check_id','object','parameter','status','priority','values_by_section',
        'explanation','sources','genplan_position','strong_evidence_count',
    ]
    document_columns = [
        'document','document_type','page_count','size_mb','status','completeness_status',
        'processing_error','document_profile',
    ]
    finding_columns = [
        'document','document_type','page','section_title','table_title','table_row',
        'parameter_code','parameter_name','object_hint','genplan_position','value_text',
        'unit','confidence','binding_status','row_integrity_status','row_integrity_reason',
        'fact_admission_decision','fact_admission_score','fact_admission_reasons',
        'evidence_quality_decision','evidence_trust_grade','physical_trace_level',
        'source_locator','project_understanding_binding','comparison_excluded',
        'engineering_plausibility_status','engineering_plausibility_reason',
        'engineering_derived_height','possible_decimal_separator_candidate',
        'match_method','structural_zone',
    ]
    object_columns = [
        'Ключ','Позиция по ГП','Наименование объекта','Статус проектирования',
        'Основание включения','Блокировка','Решение пользователя','Комментарий пользователя',
    ]
    comparison_rows=[]
    for index,row in enumerate(comparison_items,1):
        comparison_rows.append({
            'check_id':row.get('check_id') or row.get('comparison_id') or row.get('check_code') or f'XCHK-{index:03d}',
            'object':row.get('object') or row.get('object_name') or row.get('Объект') or 'Объект не определён',
            'parameter':row.get('parameter') or row.get('parameter_name') or row.get('rule_name') or row.get('Параметр') or 'Проверка',
            'status':row.get('status') or row.get('result') or row.get('Результат') or '',
            'priority':row.get('priority') or row.get('engineering_risk_level') or '',
            'values_by_section':row.get('values_by_section') or row.get('document_values') or row.get('values') or row.get('documents') or '',
            'explanation':row.get('explanation') or row.get('Пояснение') or '',
            'sources':row.get('sources') or row.get('sections') or '',
            'genplan_position':row.get('genplan_position') or row.get('Позиция по ГП') or '',
            'strong_evidence_count':row.get('strong_evidence_count') or 0,
        })
    first_doc=_first_document_record(docs)
    plausibility_rows=[{
        'Документ':row.get('document'),'Страница':row.get('page'),'Позиция':row.get('position'),
        'Объект':row.get('object'),'Заявленная высота, м':row.get('declared_height'),
        'Расчётная средняя высота, м':row.get('derived_height'),
        'Возможный десятичный кандидат, м':row.get('possible_decimal_candidate'),
        'Решение':'Требует проверки' if row.get('decision')=='HOLD' else row.get('decision'),
        'Обоснование':row.get('reason'),
    } for row in (first_doc.get('engineering_plausibility_audit') or {}).get('items') or []]
    return {
        'Тех_реестр': _excel_safe_frame(registry(docs), columns=registry_columns, max_rows=5000),
        'Тех_сверки': _excel_safe_frame(pd.DataFrame(comparison_rows), columns=comparison_columns, max_rows=10000),
        'Тех_документы': _excel_safe_frame(docs, columns=document_columns, max_rows=3000),
        'Тех_извлечение': _excel_safe_frame(engineer_findings(findings), columns=finding_columns, max_rows=10000),
        'Тех_исключённые': _excel_safe_frame(pd.DataFrame(report.get('excluded_objects') or []), columns=object_columns, max_rows=5000),
        'Тех_спорные': _excel_safe_frame(pd.DataFrame(report.get('unresolved_objects') or []), columns=object_columns, max_rows=5000),
        'Тех_правдоподобность': _excel_safe_frame(pd.DataFrame(plausibility_rows), max_rows=5000),
    }


def _safe_sheet_name(name: str) -> str:
    for ch in '[]:*?/\\':
        name = name.replace(ch, '_')
    return (name or 'Лист')[:31]


def _style_workbook(book, report_kind: str = 'gip'):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dark = '1F4E78'
    blue = 'D9EAF7'
    light = 'F3F6F9'
    green = 'E2F0D9'
    yellow = 'FFF2CC'
    red = 'FCE4D6'
    gray = 'E7E6E6'
    thin = Side(style='thin', color='B8C2CC')

    for ws in book.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30
        for cell in ws[1]:
            cell.fill = PatternFill('solid', fgColor=dark)
            cell.font = Font(color='FFFFFF', bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = Border(bottom=thin)
                value = str(cell.value or '').lower()
                if 'высок' in value or 'расхожд' in value or value == 'нет':
                    cell.fill = PatternFill('solid', fgColor=red)
                elif 'средн' in value or 'требует' in value or 'частично' in value or 'недостат' in value:
                    cell.fill = PatternFill('solid', fgColor=yellow)
                elif 'подтверж' in value or 'совпад' in value or value == 'да':
                    cell.fill = PatternFill('solid', fgColor=green)
        for idx, cells in enumerate(ws.columns, 1):
            values = [str(c.value or '') for c in cells[:120]]
            width = min(max(max((len(v) for v in values), default=0) + 2, 12), 52)
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.auto_filter.ref = ws.dimensions

    if 'Резюме' in book.sheetnames:
        ws = book['Резюме']
        ws.freeze_panes = None
        ws.auto_filter.ref = None
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 68
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill('solid', fgColor=blue)
            ws[f'B{row}'].fill = PatternFill('solid', fgColor=light)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0


def _report_context(project, docs, comparisons, risks=None, checklist_results=None, assembly_rows_data=None):
    return build_structured_report(
        project,
        docs.to_dict('records') if hasattr(docs, 'to_dict') else (docs or []),
        comparisons.to_dict('records') if hasattr(comparisons, 'to_dict') else (comparisons or []),
        risks=risks or [],
        checklist_results=checklist_results or [],
        assembly_rows=assembly_rows_data or [],
    )


def structured_excel_report(project, version, docs, findings, comparisons, *, report_kind='gip', risks=None, checklist_results=None, assembly_rows_data=None):
    doc_records=docs.to_dict('records') if hasattr(docs,'to_dict') else (docs or [])
    comparison_records=comparisons.to_dict('records') if hasattr(comparisons,'to_dict') else list(comparisons or [])
    first_record=(doc_records[0] or {}) if doc_records else {}
    register_comparison_codes={'GP_EXPLICATION_FIELD','GP_DOCUMENT_COVERAGE'}
    register_comparisons=[x for x in comparison_records if str(x.get('parameter_code') or '').upper() in register_comparison_codes]
    engineering_comparisons=[x for x in comparison_records if str(x.get('parameter_code') or '').upper() not in register_comparison_codes]
    def _cross_completed(rows):
        completed_tokens={'СОВПАДАЕТ','СООТВЕТСТВУЕТ','ПОДТВЕРЖДЕНО','ПОТЕНЦИАЛЬНОЕ РАСХОЖДЕНИЕ','РАСХОЖДЕНИЕ','КОНФЛИКТ','КОНФЛИКТ ВНУТРИ РАЗДЕЛА'}
        return sum(str(x.get('status') or x.get('result') or '').strip().upper() in completed_tokens for x in rows)
    if not checklist_results:
        checklist_results=list((first_record.get('automatic_checklist_review') or {}).get('results') or [])
    report = _report_context(project, docs, comparisons, risks, checklist_results, assembly_rows_data)
    summary = report['summary']
    out = io.BytesIO()

    # Normative validity is attached to every document by the pipeline; use the first
    # project record as the shared audit payload to avoid duplicating rows.
    normative_rows=[]
    normative_reference_details=[]
    assignment_rows=[]
    assignment_atomic_rows=[]
    assignment_summary={}
    normative_requirement_rows=[]
    normative_compliance_rows=[]
    normative_compliance_summary={}
    project_understanding={}
    project_understanding_quality={}
    if hasattr(docs,'empty') and not docs.empty:
        first_doc=docs.iloc[0].to_dict()
        normative_reference_details=list(first_doc.get('normative_validity_audit') or [])
        normative_rows=list(first_doc.get('normative_reference_summary') or normative_reference_details)
        assignment_rows=list(first_doc.get('assignment_compliance') or [])
        assignment_atomic_rows=list(first_doc.get('assignment_atomic_compliance') or [])
        assignment_summary=dict(first_doc.get('assignment_compliance_summary') or {})
        normative_requirement_rows=list(first_doc.get('normative_requirement_audit') or [])
        normative_compliance_rows=list(first_doc.get('normative_compliance_audit') or [])
        normative_compliance_summary=dict(first_doc.get('normative_compliance_summary') or {})
        project_understanding=dict(first_doc.get('project_understanding') or {})
        project_understanding_quality=dict(first_doc.get('project_understanding_quality') or {})
    elif isinstance(docs,list) and docs:
        normative_reference_details=list((docs[0] or {}).get('normative_validity_audit') or [])
        normative_rows=list((docs[0] or {}).get('normative_reference_summary') or normative_reference_details)
        assignment_rows=list((docs[0] or {}).get('assignment_compliance') or [])
        assignment_atomic_rows=list((docs[0] or {}).get('assignment_atomic_compliance') or [])
        assignment_summary=dict((docs[0] or {}).get('assignment_compliance_summary') or {})
        normative_requirement_rows=list((docs[0] or {}).get('normative_requirement_audit') or [])
        normative_compliance_rows=list((docs[0] or {}).get('normative_compliance_audit') or [])
        normative_compliance_summary=dict((docs[0] or {}).get('normative_compliance_summary') or {})
        project_understanding=dict((docs[0] or {}).get('project_understanding') or {})
        project_understanding_quality=dict((docs[0] or {}).get('project_understanding_quality') or {})
    # Prefer the plan produced after Deep Evidence adjudication.  Rebuilding it
    # from legacy statuses would discard adversarial downgrades.
    review_plan=dict(first_record.get('project_review_plan') or {})
    if not review_plan:
        review_plan=build_review_plan(doc_records,checklist_results or [])
    review_domains=review_plan.get('domains') or {}
    coverage_matrix_payload=dict(first_record.get('coverage_matrix') or {})
    semantic_engine_summary=dict(first_record.get('semantic_evidence_engine') or {})
    ai_summary_rows=[]
    ai_execution_rows=[]
    ai_call_rows=[]
    for domain_key, domain_label in (('assignment','Задание на проектирование'),('checklist','Чек-листы')):
        audit=dict(semantic_engine_summary.get(domain_key) or {})
        if not audit:
            continue
        preflight=dict(audit.get('preflight') or {})
        judge_preflight=dict(preflight.get('judge') or {})
        critic_preflight=dict(preflight.get('critic') or {})
        ai_summary_rows.append({
            'Контур':domain_label,
            'Режим включён':'Да' if audit.get('enabled') else 'Нет',
            'Настроенная проверяющая модель':audit.get('configured_judge_provider') or 'Не настроен',
            'Настроенная контрольная модель':audit.get('configured_critic_provider') or 'Не настроен',
            'Готово атомарных пакетов L4':audit.get('judge_candidates',0),
            'Отобрано для проверяющей модели':audit.get('judge_selected',0),
            'Предварительная проверка основной модели':ru_label(judge_preflight.get('state') or 'NOT_REQUIRED'),
            'Фактический основной провайдер':judge_preflight.get('actual_provider') or '—',
            'Основная модель':judge_preflight.get('model') or '—',
            'Предварительная проверка контрольной модели':ru_label(critic_preflight.get('state') or 'NOT_REQUIRED'),
            'Фактический контрольный провайдер':critic_preflight.get('actual_provider') or '—',
            'Контрольная модель':critic_preflight.get('model') or '—',
            'Ответов проверяющей модели':audit.get('judge_responses',0),
            'Ответов контрольной модели':audit.get('critic_responses',0),
            'Подтверждено консенсусом':audit.get('promoted_verified',0),
            'Подтверждено несоответствий':audit.get('project_findings',0),
            'Заблокировано контролем':audit.get('blocked_consensus',0),
            'Причина неактивности':' | '.join(audit.get('activation_reasons') or []),
            'Ошибки проверяющей модели — кратко':_ai_error_summary(audit.get('judge_errors') or []),
            'Ошибки контрольной модели — кратко':_ai_error_summary(audit.get('critic_errors') or []),
        })
        for role_key, role_label in (('judge','Предварительная проверка основной модели'),('critic','Предварительная проверка контрольной модели')):
            check=dict(preflight.get(role_key) or {})
            if not check:
                continue
            ai_call_rows.append({
                'Контур':domain_label, 'Роль':role_label, 'Попытка':0,
                'Настроенный провайдер':check.get('configured_provider') or '—',
                'Фактический провайдер':check.get('actual_provider') or '—',
                'Модель':check.get('model') or '—', 'Код HTTP':check.get('status_code'),
                'Запрошено пакетов':0, 'Получено ответов':0,
                'Состояние':ru_label(check.get('state')), 'Ошибка':str(check.get('error') or '')[:1200],
                'ID пакетов':'',
            })
        for row in audit.get('execution_log') or []:
            ai_execution_rows.append({
                'Контур':domain_label,
                'ID пакета':row.get('packet_id'),
                'Семейство проверки':ru_label(row.get('checker_family')),
                'Уровень доказательства':ru_label(row.get('evidence_level')),
                'Контракт доказательства':ru_label(row.get('evidence_contract_state')),
                'Отобран':'Да' if row.get('selected') else 'Нет',
                'Основание отбора':row.get('selection_reason'),
                'Решение проверяющей модели':ru_label(row.get('judge_state')),
                'Ответ проверяющей модели получен':'Да' if row.get('judge_response_received') else 'Нет',
                'Решение прошло структурный контроль':'Да' if row.get('judge_valid') else 'Нет',
                'Достоверность проверяющей модели':row.get('judge_confidence'),
                'Фактический основной провайдер':row.get('judge_provider') or '—',
                'Основная модель':row.get('judge_model') or '—',
                'Решение контрольной модели':ru_label(row.get('critic_state')),
                'Ответ контрольной модели получен':'Да' if row.get('critic_response_received') else 'Нет',
                'Фактический контрольный провайдер':row.get('critic_provider') or '—',
                'Контрольная модель':row.get('critic_model') or '—',
                'Итог консенсуса':ru_label(row.get('consensus_state')),
                'Причины блокировки':' | '.join(row.get('blocking_reasons') or []),
            })
        for role_key, role_label in (('judge_calls','Проверяющая модель'),('critic_calls','Контрольная модель')):
            for call in audit.get(role_key) or []:
                ai_call_rows.append({
                    'Контур':domain_label,
                    'Роль':role_label,
                    'Попытка':call.get('attempt'),
                    'Настроенный провайдер':call.get('configured_provider') or '—',
                    'Фактический провайдер':call.get('actual_provider') or '—',
                    'Модель':call.get('model') or '—',
                    'Код HTTP':call.get('status_code'),
                    'Запрошено пакетов':call.get('requested',0),
                    'Получено ответов':call.get('responses',0),
                    'Состояние':ru_label(call.get('state')),
                    'Ошибка':str(call.get('error') or '')[:1200],
                    'ID пакетов':' | '.join(call.get('packet_ids') or []),
                })
    ai_summary_df=_excel_safe_frame(pd.DataFrame(ai_summary_rows))
    ai_execution_df=_excel_safe_frame(pd.DataFrame(ai_execution_rows))
    ai_calls_df=_excel_safe_frame(pd.DataFrame(ai_call_rows))
    normative_statuses={}
    for row in normative_rows:
        status=str(row.get('status') or 'Требует верификации')
        normative_statuses[status]=normative_statuses.get(status,0)+1
    normative_attention=sum(v for k,v in normative_statuses.items() if k not in {'Действует','Действует с изменениями'})
    normative_high=sum(1 for x in normative_rows if x.get('status') in {'Заменён','Утратил силу'} and x.get('impact_risk')=='Высокий')

    assignment_plan=review_domains.get('Задание на проектирование',{})
    normative_plan=review_domains.get('НТД',{})
    checklist_plan=review_domains.get('Чек-листы',{})
    report_quality_gate=dict(first_record.get('report_quality_gate') or {})
    normative_registry_verified=sum(1 for x in normative_rows if x.get('coverage_status')=='Проверено по реестру')
    total_project_findings=int(summary.get('project_findings',0) or 0)+int(review_plan.get('project_findings',0) or 0)
    total_review_questions=int(summary.get('review_questions',0) or 0)+int(review_plan.get('review_questions',0) or 0)
    total_system_limitations=int(summary.get('system_limitations',0) or 0)+int(review_plan.get('system_limitations',0) or 0)
    conclusion_parts=[]
    if total_project_findings:
        conclusion_parts.append(f"Подтверждённых несоответствий проекта: {total_project_findings}.")
    if total_review_questions:
        conclusion_parts.append(f"Вопросов, требующих проверки специалистом: {total_review_questions}.")
    if total_system_limitations:
        conclusion_parts.append(f"Проверок вне текущего автоматического покрытия: {total_system_limitations}.")
    final_conclusion=' '.join(conclusion_parts) or report['conclusion']
    report_status='Итоговый' if summary.get('completeness')=='Подтверждена' else 'Предварительный — состав/комплектность проекта не подтверждены пользователем'
    if report_status.startswith('Предварительный'):
        final_conclusion='Предварительный результат: окончательный вывод удержан до подтверждения состава и комплектности проекта. '+final_conclusion
    summary_rows = [
        ['Наименование проекта', project],
        ['Дата и время проверки', datetime.now().strftime('%d.%m.%Y %H:%M')],
        ['Версия ExpertCheck', version],
        ['Статус отчёта', report_status],
        ['Комплектность', summary['completeness']],
        ['Загружено документов', summary['documents']],
        ['Подтверждено объектов', summary.get('objects_confirmed',summary['objects'])],
        ['Объектов требуют подтверждения источника', summary.get('objects_unresolved',0)],
        ['Подтверждённых несоответствий проекта', total_project_findings],
        ['Вопросов специалисту', total_review_questions],
        ['Проверок вне автоматического покрытия', total_system_limitations],
        ['Строгое покрытие L5, %', coverage_matrix_payload.get('coverage_pct',0)],
        ['Доказательное покрытие L3–L5, %', coverage_matrix_payload.get('evidence_coverage_pct',0)],
        ['Проверок с независимым AI-консенсусом', coverage_matrix_payload.get('semantic_consensus_completed',0)],
        ['Задание: покрытие автоматической проверки, %', assignment_plan.get('coverage_pct',0)],
        ['Задание: доказательное покрытие L3–L5, %', assignment_plan.get('evidence_coverage_pct',0)],
        ['Задание: подтверждено', assignment_plan.get('confirmed',0)],
        ['Задание: выявлено несоответствий', assignment_plan.get('issue',0)],
        ['НТД: покрытие доказательной проверки, %', normative_plan.get('coverage_pct',0)],
        ['НТД: подтверждено требований', normative_plan.get('confirmed',0)],
        ['НТД: выявлено несоответствий', normative_plan.get('issue',0)],
        ['НТД: обнаружено уникальных ссылок', len(normative_rows)],
        ['НТД: проверено по реестру актуальности', normative_registry_verified],
        ['Чек-листы: покрытие автоматической проверки, %', checklist_plan.get('coverage_pct',0)],
        ['Чек-листы: доказательное покрытие L3–L5, %', checklist_plan.get('evidence_coverage_pct',0)],
        ['Чек-листы: подтверждено', checklist_plan.get('confirmed',0)],
        ['Чек-листы: выявлено несоответствий', checklist_plan.get('issue',0)],
        ['Сверка реестров/чертежей: завершено', f"{_cross_completed(register_comparisons)} из {len(register_comparisons)}"],
        ['Инженерные параметры: завершено', f"{_cross_completed(engineering_comparisons)} из {len(engineering_comparisons)}"],
        ['Контроль согласованности отчёта', 'Пройден' if report_quality_gate.get('status')=='PASSED' else 'Требует проверки'],
        ['Итоговый вывод', final_conclusion],
    ]
    if report_kind=='manager':
        manager_metrics={
            'Наименование проекта','Дата и время проверки','Версия ExpertCheck','Статус отчёта','Комплектность',
            'Загружено документов','Подтверждено объектов','Объектов требуют подтверждения источника','Подтверждённых несоответствий проекта',
            'Вопросов специалисту','Задание: покрытие автоматической проверки, %',
            'Проверок вне автоматического покрытия',
            'Строгое покрытие L5, %','Доказательное покрытие L3–L5, %','Проверок с независимым AI-консенсусом',
            'Задание: доказательное покрытие L3–L5, %','Чек-листы: доказательное покрытие L3–L5, %',
            'НТД: покрытие доказательной проверки, %','Чек-листы: покрытие автоматической проверки, %',
            'Сверка реестров/чертежей: завершено','Инженерные параметры: завершено',
            'Контроль согласованности отчёта','Итоговый вывод',
        }
        summary_rows=[row for row in summary_rows if row[0] in manager_metrics]
    elif report_kind=='gip':
        # The GIP view keeps decision metrics but omits developer-oriented NTD
        # inventory detail, which remains in the technical appendix.
        summary_rows=[row for row in summary_rows if row[0] not in {
            'НТД: обнаружено уникальных ссылок','НТД: проверено по реестру актуальности'
        }]

    selected_risks = [r for r in report['risks'] if r.get('level') in ({'Высокий','Средний'} if report_kind != 'technical' else {'Высокий','Средний','Низкий'})]
    if report_kind == 'manager': selected_risks = selected_risks[:12]
    elif report_kind == 'gip': selected_risks = selected_risks[:30]
    risks_df = pd.DataFrame([{
        'ID': r.get('risk_id'),
        'Уровень': r.get('level'),
        'Категория': r.get('category'),
        'Объект / раздел': r.get('object') or '—',
        'Вопрос': r.get('parameter'),
        'Выявленная проблема': r.get('finding'),
        'Возможное замечание': r.get('possible_remark'),
        'Рекомендуемое действие': r.get('recommendation'),
        'Источники': r.get('sources'),
        'Сценарий базы': r.get('scenario_id'),
        'Повторяемость': r.get('recurrence'),
        'Проекты-аналоги': ', '.join(r.get('analog_projects') or []),
        'Решение пользователя': r.get('user_status') or 'Не рассмотрено',
        'Комментарий пользователя': r.get('user_comment') or '',
    } for r in selected_risks])

    atomic_problems=[{
        'id':x.get('atom_id') or x.get('requirement_id'),
        'object':x.get('object_name') or 'Задание на проектирование',
        'parameter':x.get('atom_text') or x.get('requirement_text'),
        'status':'Выявлено несоответствие',
        'priority':'Высокий' if x.get('atomic_kind') in {'VALUE_COMPARISON','PROHIBITION','EQUIPMENT_IDENTITY'} else 'Средний',
        'values':x.get('difference') or '',
        'explanation':x.get('decision_basis') or '',
        'sources':_evidence_sources(x),
        'finding_type':'PROJECT_FINDING',
    } for x in assignment_atomic_rows if str(x.get('verification_kind') or '')=='PROJECT_FINDING']
    all_problems=list(report['problems'] or [])+atomic_problems
    confirmed_problems=[x for x in all_problems if str(x.get('finding_type') or '').upper()=='PROJECT_FINDING']
    selected_problems = confirmed_problems if report_kind == 'technical' else confirmed_problems[:(12 if report_kind == 'manager' else 35)]
    problems_df = pd.DataFrame(selected_problems).rename(columns={
        'id':'ID', 'object':'Объект', 'parameter':'Показатель', 'status':'Результат',
        'priority':'Приоритет', 'values':'Значения по разделам', 'explanation':'Пояснение', 'sources':'Источники',
    })
    question_rows=[]
    for item in all_problems:
        if str(item.get('finding_type') or '').upper()!='REVIEW_QUESTION':
            continue
        question_rows.append({
            'ID':item.get('id'),'Контур':'Межраздельная сверка','Объект':item.get('object') or '—',
            'Проверка':item.get('parameter'),'Причина':item.get('explanation') or '',
            'Недостающие доказательства':'Уточнить доверенные источники и актуальность разделов',
            'Ожидаемые разделы':item.get('sources') or '—','Уровень доказательства':'—',
        })
    for item in review_plan.get('items') or []:
        if str(item.get('verification_kind') or '').upper()!='REVIEW_QUESTION':
            continue
        question_rows.append({
            'ID':item.get('plan_id'),'Контур':item.get('domain'),'Объект':item.get('entity') or '—',
            'Проверка':item.get('title'),'Причина':item.get('coverage_reason') or 'Требуется предметное решение специалиста.',
            'Недостающие доказательства':', '.join(ru_label(v) for v in (item.get('missing_evidence_slots') or [])) or '—',
            'Ожидаемые разделы':', '.join(item.get('expected_evidence_route') or item.get('expected_sections') or []) or '—',
            'Уровень доказательства':ru_label(item.get('evidence_level') or 'L0'),
        })
    deduped_questions=[]; seen_questions=set()
    for item in question_rows:
        key=(str(item.get('Контур') or ''),str(item.get('ID') or ''),str(item.get('Проверка') or ''))
        if key in seen_questions:
            continue
        seen_questions.add(key); deduped_questions.append(item)
    selected_questions=deduped_questions[:40] if report_kind=='manager' else deduped_questions
    questions_df=pd.DataFrame(selected_questions)
    object_df = pd.DataFrame(report['confirmed_objects']).rename(columns={
        'position':'Поз.', 'name':'Наименование объекта', 'status':'Статус', 'source':'Основной источник',
    })
    checklist_all_df = pd.DataFrame([{
        'Раздел': r.get('automatic_section') or r.get('section') or r.get('Раздел') or 'Не определён',
        'Чек-лист': r.get('automatic_checklist') or r.get('source_file') or r.get('Чек-лист') or '',
        'Пункт': f"{r.get('item_no') or r.get('position') or ''} — {r.get('question') or r.get('Позиция по чек-листу') or ''}".strip(' —'),
        'Результат ExpertCheck': r.get('status') or r.get('Соответствие') or r.get('result'),
        'Уровень проверки': r.get('verification_level') or r.get('check_level') or r.get('typed_check') or r.get('execution_class') or '—',
        'Решение специалиста': r.get('specialist_decision') or 'Не рассмотрено',
        'Комментарий специалиста': r.get('specialist_comment') or '',
        'Обоснование ExpertCheck': r.get('evidence') or r.get('Обоснование') or '',
        'Итоговый класс проверки': verification_label(r.get('final_verification_kind') or r.get('verification_kind')),
        'Архетип покрытия':ru_label(r.get('coverage_archetype')),
        'Код причины незавершения':r.get('coverage_reason_code') or '',
        'Причина незавершения':r.get('coverage_reason') or '',
        'Недостающие слоты':', '.join(r.get('missing_evidence_slots') or []),
        'Углублённый анализ доказательств': ru_label(r.get('deep_evidence_state')),
        'Уровень доказательства':ru_label(r.get('evidence_level') or 'L0'),
        'Доказательное покрытие атомов, %':r.get('evidence_coverage_pct'),
        'Смысловой консенсус':ru_label(r.get('semantic_consensus_state')),
        'Смысловых условий завершено':r.get('semantic_consensus_completed',0),
        'Причины ограничения': ' | '.join(r.get('deep_evidence_reasons') or []),
        'Источники': _evidence_sources(r),
    } for r in report['checklist_results'] if not r.get('is_heading')])

    checklist_problem_df = pd.DataFrame([{
        'Раздел': r.get('automatic_section') or r.get('section') or r.get('Раздел') or 'Не определён',
        'Чек-лист': r.get('automatic_checklist') or r.get('source_file') or r.get('Чек-лист') or '',
        'Пункт': f"{r.get('item_no') or r.get('position') or ''} — {r.get('question') or r.get('Позиция по чек-листу') or ''}".strip(' —'),
        'Результат': r.get('status') or r.get('Соответствие') or r.get('result'),
        'Обоснование': r.get('evidence') or r.get('Обоснование') or '',
        'Источники': _evidence_sources(r),
    } for r in report['checklist_results'] if str(r.get('status') or r.get('Соответствие') or r.get('result') or '').lower() in {'нет','частично','требует проверки','нет данных','не соответствует'}])
    actions=list(report['recommendations'] or [])
    def add_action(text):
        if text and text not in actions: actions.append(text)
    for label,domain in (
        ('Задание на проектирование',assignment_plan),('Требования НТД',normative_plan),('Чек-листы',checklist_plan)
    ):
        pending=int(domain.get('review',0) or 0)
        if pending:
            add_action(f"{label}: рассмотреть {pending} адресных вопросов специалиста и зафиксировать решение.")
    for atom in assignment_atomic_rows:
        if str(atom.get('verification_kind') or '')=='PROJECT_FINDING':
            add_action(
                f"Задание, {atom.get('atom_id') or atom.get('requirement_id')}: "
                f"{atom.get('recommendation') or 'устранить подтверждённое отклонение либо согласовать изменение Задания.'}"
            )
    comparison_pending=int(summary.get('review_questions',0) or 0)
    if comparison_pending:
        add_action(f"Межраздельная сверка: получить недостающие доказательства по {comparison_pending} проверкам.")
    if normative_attention:
        add_action(f"Ссылки НТД: проверить актуальность и редакции {normative_attention} позиций.")
    recommendations_df = pd.DataFrame({'Приоритетное действие': actions})
    review_plan_df = pd.DataFrame([{
        'Контур проверки':x.get('domain'),
        'Проверка':x.get('title'),
        'Результат':verification_label(x.get('status')),
        'Область':ru_label(x.get('scope')) if x.get('scope') else '—',
        'Тип проверки':ru_label(x.get('check_type')) if x.get('check_type') else '—',
        'Архетип покрытия':ru_label(x.get('coverage_archetype')) or '—',
        'Состояние покрытия':ru_label(x.get('coverage_state')) or '—',
        'Уровень доказательства':ru_label(x.get('evidence_level') or 'L0'),
        'Доказательное покрытие атомов, %':x.get('evidence_coverage_pct'),
        'Семейство проверяющего механизма':ru_label(x.get('checker_family')) if x.get('checker_family') else '—',
        'Режим проверяющего механизма':ru_label(x.get('checker_mode')) if x.get('checker_mode') else '—',
        'Независимый смысловой консенсус':ru_label(x.get('semantic_consensus_state')) if x.get('semantic_consensus_state') else '—',
        'Ожидаемое доказательство':x.get('expected_evidence') or '',
        'Маршрут доказательства':', '.join(x.get('expected_evidence_route') or x.get('expected_sections') or []),
        'Итоговый класс':verification_label(x.get('verification_kind')),
        'Проверка достаточности':ru_label(x.get('adversarial_state') or x.get('deep_evidence_state')),
        'Код причины незавершения':ru_label(x.get('coverage_reason_code')) if x.get('coverage_reason_code') else '',
        'Причина незавершения':x.get('coverage_reason') or ' | '.join(x.get('adversarial_reasons') or x.get('deep_evidence_reasons') or []),
        'Недостающие слоты':', '.join(ru_label(v) for v in (x.get('missing_evidence_slots') or [])),
    } for x in review_plan.get('items') or []])
    limitations_df=pd.DataFrame([{
        'Контур':x.get('domain'),'Проверка':x.get('title'),'Архетип':ru_label(x.get('coverage_archetype')),
        'Уровень доказательства':ru_label(x.get('evidence_level') or 'L0'),
        'Код ограничения':ru_label(x.get('coverage_reason_code') or 'UNSPECIFIED'),
        'Что не получено':x.get('coverage_reason') or 'Доказательный контракт не завершён.',
        'Недостающие слоты':', '.join(ru_label(v) for v in (x.get('missing_evidence_slots') or [])),
        'Ожидаемые разделы':', '.join(x.get('expected_evidence_route') or x.get('expected_sections') or []),
        'Статус':'Ограничение ExpertCheck — не замечание проекта',
    } for x in review_plan.get('items') or [] if str(x.get('verification_kind') or '').upper()=='SYSTEM_LIMITATION'])
    coverage_matrix_df=pd.DataFrame([{
        'Архетип':ru_label(x.get('archetype')),'Всего':x.get('total',0),'Завершено автоматически':x.get('completed',0),
        'Покрытие, %':x.get('coverage_pct',0),'Подтверждено':x.get('verified_ok',0),
        'Доказательства готовы L3–L5':x.get('evidence_ready',0),'Доказательное покрытие, %':x.get('evidence_coverage_pct',0),
        'Независимый AI-консенсус':x.get('semantic_consensus_completed',0),
        'L0':(x.get('evidence_levels') or {}).get('L0',0),'L1':(x.get('evidence_levels') or {}).get('L1',0),
        'L2':(x.get('evidence_levels') or {}).get('L2',0),'L3':(x.get('evidence_levels') or {}).get('L3',0),
        'L4':(x.get('evidence_levels') or {}).get('L4',0),'L5':(x.get('evidence_levels') or {}).get('L5',0),
        'Несоответствия':x.get('project_findings',0),'Вопросы специалисту':x.get('review_questions',0),
        'Ограничения системы':x.get('system_limitations',0),'Доверенные рецепты':x.get('trusted_recipes',0),
        'Только поиск кандидатов':x.get('retrieval_only_recipes',0),
        'Основные причины пробелов':' | '.join(f"{ru_label(r.get('code'))}: {r.get('count')}" for r in (x.get('top_gap_reasons') or [])),
    } for x in coverage_matrix_payload.get('matrix') or []])
    understanding_rows=[]
    for obj in project_understanding.get('objects') or []:
        for prop in obj.get('property_summary') or []:
            paired_values='; '.join(
                f"{item.get('section')}: {' | '.join(str(value) for value in item.get('values') or [])}"
                for item in prop.get('values_by_section') or []
            )
            understanding_rows.append({
              'ID объекта':obj.get('object_id'),'Объект':obj.get('name'),'Позиция по ГП':obj.get('position') or '—',
              'Тип объекта':obj.get('object_type'),'Показатель':prop.get('parameter_name'),
              'Код показателя':parameter_label(prop.get('parameter_code')),'Разделы':', '.join(prop.get('sections') or []),
              'Количество доказательств':prop.get('evidence_count',0),
              'Профильный источник':'Да' if prop.get('owner_evidence') else 'Нет',
              'Конфликт значений':'Да' if prop.get('value_conflict') else 'Нет',
              'Значения по разделам':paired_values or ' | '.join(str(x) for x in prop.get('values') or []),
            })
    understanding_df=pd.DataFrame(understanding_rows)

    assignment_df = pd.DataFrame([{
        'ID требования':x.get('requirement_id'),
        'Строка Задания':x.get('source_row') or '—',
        'Раздел / вопрос Задания':x.get('source_row_title') or '—',
        'Тип проверки':ru_label(x.get('requirement_type')),
        'Область требования':ru_label(x.get('requirement_scope') or (x.get('evidence_contract_v2') or {}).get('scope')),
        'Метод проверки':ru_label((x.get('evidence_contract_v2') or {}).get('check_method')),
        'Ожидаемые разделы':', '.join((x.get('evidence_contract_v2') or {}).get('expected_sections') or []),
        'Способ восстановления':ru_label(x.get('cell_reconstruction')),
        'Требование Задания':x.get('requirement_text'),
        'Объект':x.get('object_name') or '—',
        'Показатель':parameter_label(x.get('parameter_code')) if x.get('parameter_code') else '—',
        'Требуемое значение':x.get('required_value'),
        'Ед. изм.':x.get('unit'),
        'Результат':x.get('status'),
        'Документ':x.get('source_document'),
        'Страница':x.get('page'),
        'Доказательства':' | '.join(x.get('evidence') or []),
        'Качество доказательства':ru_label(x.get('evidence_quality_state')),
        'Направленных кандидатов':len(x.get('directed_evidence_candidates') or []),
        'Ожидаемое доказательство':x.get('expected_evidence') or '',
        'Основание вывода':x.get('decision_basis') or '',
        'Рекомендация':x.get('recommendation'),
        'Итоговый класс проверки':verification_label(x.get('final_verification_kind') or x.get('verification_kind')),
        'Архетип покрытия':ru_label(x.get('coverage_archetype')),
        'Код причины незавершения':ru_label(x.get('coverage_reason_code')) if x.get('coverage_reason_code') else '',
        'Причина незавершения':x.get('coverage_reason') or '',
        'Недостающие слоты':', '.join(ru_label(v) for v in (x.get('missing_evidence_slots') or [])),
        'Углублённый анализ доказательств':ru_label(x.get('deep_evidence_state')),
        'Уровень доказательства':ru_label(x.get('evidence_level') or 'L0'),
        'Доказательное покрытие атомов, %':x.get('evidence_coverage_pct'),
        'Смысловой консенсус':ru_label(x.get('semantic_consensus_state')),
        'Причины ограничения':' | '.join(x.get('deep_evidence_reasons') or []),
    } for x in assignment_rows])
    assignment_gip_df=_excel_safe_frame(assignment_df,columns=[
        'ID требования','Раздел / вопрос Задания','Требование Задания','Объект','Показатель',
        'Требуемое значение','Ед. изм.','Результат','Документ','Страница','Доказательства',
        'Основание вывода','Рекомендация','Итоговый класс проверки','Архетип покрытия',
        'Код причины незавершения','Причина незавершения','Недостающие слоты','Углублённый анализ доказательств','Причины ограничения',
    ])
    assignment_atomic_df=pd.DataFrame([{
        'ID атомарного условия':x.get('atom_id') or x.get('requirement_id'),
        'ID исходного требования':x.get('parent_requirement_id'),
        'Строка Задания':x.get('source_row') or '—',
        'Атомарное условие':x.get('atom_text') or x.get('requirement_text'),
        'Тип атома':ru_label(x.get('atomic_kind')),
        'Объект':x.get('object_name') or '—',
        'Показатель':parameter_label(x.get('parameter_code')) if x.get('parameter_code') else x.get('focus') or '—',
        'Оператор условия':ru_label(x.get('comparison_operator') or _difference_field(x,'operator')),
        'Требуемое значение':x.get('required_value'),
        'Нижняя граница':x.get('required_min'),
        'Верхняя граница':x.get('required_max'),
        'Ед. изм.':x.get('unit'),
        'Нормализованное требуемое значение':_difference_field(x,'required'),
        'Нормализованная нижняя граница':_difference_field(x,'required_min'),
        'Нормализованная верхняя граница':_difference_field(x,'required_max'),
        'Значение проекта':_difference_field(x,'observed'),
        'Единица сравнения':_difference_field(x,'unit'),
        'Условие выполнено':('Да' if _difference_field(x,'satisfied',None) is True else 'Нет' if _difference_field(x,'satisfied',None) is False else '—'),
        'Ожидаемые разделы':', '.join((x.get('verification_recipe') or {}).get('expected_sections') or x.get('expected_sections') or []),
        'Рецепт':x.get('recipe_id'),
        'Статус рецепта':ru_label(x.get('recipe_status')),
        'Результат':x.get('status'),
        'Итоговый класс':verification_label(x.get('final_verification_kind') or x.get('verification_kind')),
        'Доказательства':' | '.join(x.get('evidence') or []),
        'Явное различие':x.get('difference') or '',
        'Основание вывода':x.get('decision_basis') or '',
        'Рекомендация':x.get('recommendation') or '',
        'Контроль рецепта':ru_label(x.get('critic_state')),
        'Регрессионный контроль':ru_label(x.get('regression_state')),
        'Контракт доказательства':ru_label(x.get('evidence_contract_state')),
        'Архетип покрытия':ru_label(x.get('coverage_archetype')),
        'Состояние покрытия':ru_label(x.get('coverage_state')),
        'Код причины незавершения':ru_label(x.get('coverage_reason_code')) if x.get('coverage_reason_code') else '',
        'Причина незавершения':x.get('coverage_reason') or '',
        'Недостающие слоты':', '.join(ru_label(v) for v in (x.get('missing_evidence_slots') or [])),
        'Смысловой контроль':ru_label(x.get('semantic_gate_state')),
        'Требуемая модальность':ru_label((x.get('evidence_contract') or x.get('evidence_contract_v2') or {}).get('required_modality')),
        'Критические квалификаторы':', '.join((x.get('evidence_contract') or x.get('evidence_contract_v2') or {}).get('critical_qualifiers') or []),
        'Причины смыслового удержания':' | '.join(x.get('semantic_gate_reasons') or []),
        'Уровень доказательства':ru_label(x.get('evidence_level') or 'L0'),
        'Причина уровня доказательства':x.get('evidence_level_reason') or '',
        'Семейство проверяющего механизма':ru_label(x.get('checker_family')),
        'Режим проверяющего механизма':ru_label(x.get('checker_mode')),
        'Смысловой консенсус':ru_label(x.get('semantic_consensus_state')),
        'Решение проверяющей модели':ru_label((x.get('semantic_judge') or {}).get('verdict')),
        'Достоверность проверяющей модели':(x.get('semantic_judge') or {}).get('confidence'),
        'Провайдер проверяющей модели':(x.get('semantic_judge') or {}).get('provider'),
        'Контрольная модель приняла':('Да' if (x.get('semantic_critic') or {}).get('accept') is True else 'Нет' if x.get('semantic_critic') else '—'),
        'Провайдер контрольной модели':(x.get('semantic_critic') or {}).get('provider'),
        'Причины блокировки консенсуса':' | '.join(x.get('semantic_consensus_reasons') or []),
        'Углублённый анализ доказательств':ru_label(x.get('deep_evidence_state') or x.get('adversarial_state')),
        'Причины ограничения':' | '.join(x.get('deep_evidence_reasons') or x.get('adversarial_reasons') or []),
    } for x in assignment_atomic_rows])

    normative_requirement_df = pd.DataFrame([{
        'НТД':x.get('reference'),
        'Пункт / статья':x.get('clause') or '—',
        'Тип требования':x.get('modality'),
        'Тип проверки':ru_label(x.get('check_kind')) or '—',
        'Качество нормативного основания':x.get('requirement_quality') or '—',
        'Контекст в проекте':x.get('project_context'),
        'Статус НТД':x.get('normative_status'),
        'Статус редакции':x.get('edition_status'),
        'Структурированное требование':x.get('curated_requirement') or 'Не загружено',
        'Статус анализа требования':x.get('analysis_status'),
        'Файл':x.get('document'),
        'Страница':x.get('page'),
        'Риск влияния':x.get('impact_risk'),
    } for x in normative_requirement_rows])

    normative_df = pd.DataFrame([{
        'НТД':x.get('reference'),
        'Статус':x.get('status'),
        'Статус редакции':(x.get('edition_assessment') or {}).get('edition_status',''),
        'Актуальная редакция / замена':(x.get('edition_assessment') or {}).get('current_reference','') or x.get('replacement',''),
        'Количество упоминаний':x.get('mentions',1),
        'Документов':x.get('documents_count',1 if x.get('document') else 0),
        'Где встречается':'; '.join(x.get('pages') or []) if isinstance(x.get('pages'),list) else f"{x.get('document') or ''}, стр. {x.get('page') or ''}",
        'Риск влияния':x.get('impact_risk') if x.get('coverage_status')!='Требует наполнения KB' else 'Не оценивается — пробел KB',
        'Покрытие ExpertCheck':x.get('coverage_status') or '—',
        'Приоритет базы':x.get('verification_priority'),
        'Дата проверки':x.get('verified_on') or x.get('last_verified_at') or '',
        'Источник проверки':x.get('official_source'),
        'Рекомендуемый официальный источник':x.get('official_source_candidate') or '',
    } for x in normative_rows])

    summary_df = _excel_safe_frame(pd.DataFrame(summary_rows, columns=['Показатель', 'Значение']))
    risks_df = _excel_safe_frame(risks_df)
    problems_df = _excel_safe_frame(problems_df)
    questions_df = _excel_safe_frame(questions_df)
    object_df = _excel_safe_frame(object_df)
    checklist_problem_df = _excel_safe_frame(checklist_problem_df)
    checklist_all_df = _excel_safe_frame(checklist_all_df)
    recommendations_df = _excel_safe_frame(recommendations_df)
    review_plan_df = _excel_safe_frame(review_plan_df)
    limitations_df = _excel_safe_frame(limitations_df)
    coverage_matrix_df = _excel_safe_frame(coverage_matrix_df)

    normative_compliance_df = pd.DataFrame([{
        'ID требования':x.get('requirement_id'),
        'Тип знания':ru_label(x.get('knowledge_kind') or 'LAW_REQUIREMENT'),
        'НТД':x.get('source'),
        'Пункт / статья':x.get('paragraph') or '—',
        'Тема':x.get('topic'),
        'Требование':x.get('requirement'),
        'Тип проверки':ru_label(x.get('check_kind')),
        'Пункт верифицирован':'Да' if x.get('verified_clause') else 'Нет',
        'Готово к смысловой AI-проверке':'Да' if x.get('ai_review_ready') else 'Нет',
        'Результат':x.get('status'),
        'Покрытие':ru_label(x.get('coverage_state')),
        'Основание вывода':x.get('decision_basis'),
        'Доказательства':' | '.join(f"{e.get('document')}, стр. {e.get('page')}: {e.get('context') or e.get('text') or e.get('value') or ''}" for e in (x.get('evidence') or [])[:5]),
    } for x in normative_compliance_rows])
    normative_df = _excel_safe_frame(normative_df)
    assignment_df = _excel_safe_frame(assignment_df)
    assignment_atomic_df = _excel_safe_frame(assignment_atomic_df)
    understanding_df = _excel_safe_frame(understanding_df)
    normative_requirement_df = _excel_safe_frame(normative_requirement_df)
    normative_compliance_df = _excel_safe_frame(normative_compliance_df)

    sheets: list[tuple[str, pd.DataFrame]] = [('Резюме', summary_df)]
    # Подтверждённые расхождения и адресные вопросы имеют разный доказательный
    # статус и поэтому всегда выводятся на отдельных листах.
    if not problems_df.empty:
        if report_kind == 'manager':
            sheets.append(('Подтверждённые расхождения', _excel_safe_frame(problems_df, columns=[
                'ID','Объект','Показатель','Результат','Приоритет','Пояснение','Источники',
            ])))
        else:
            sheets.append(('Подтверждённые расхождения', problems_df))
    if not questions_df.empty:
        sheets.append(('Вопросы специалисту', questions_df))
    if report_kind == 'manager':
        def readiness_row(label,domain):
            total=int(domain.get('total',0) or 0); ok=int(domain.get('confirmed',0) or 0); issues=int(domain.get('issue',0) or 0)
            attention=int(domain.get('review',0) or 0)+int(domain.get('system_limitation',0) or 0)
            return {'Контур':label,'Всего':total,'Завершено автоматически':ok+issues,'Подтверждённые несоответствия':issues,'Требует внимания':attention,'Строгое покрытие L5, %':round(100*(ok+issues)/max(1,total),1),'Доказательное покрытие L3–L5, %':domain.get('evidence_coverage_pct',0),'Независимый AI-консенсус':domain.get('semantic_consensus_completed',0)}
        normative_valid_verified=sum(1 for x in normative_rows if x.get('coverage_status')=='Проверено по реестру' and x.get('status') in {'Действует','Действует с изменениями'})
        def comparison_count(rows,finding_type):
            return sum(
                str(classify_finding(x,source_kind='comparison').get('finding_type') or '').upper()==finding_type
                for x in rows or []
            )
        cross_total=len(comparison_records)
        cross_ok=sum(1 for x in comparison_records if str(x.get('status') or x.get('result') or '').strip().upper() in {'СОВПАДАЕТ','СООТВЕТСТВУЕТ','ПОДТВЕРЖДЕНО'})
        cross_issues=comparison_count(comparison_records,'PROJECT_FINDING')
        cross_attention=comparison_count(comparison_records,'REVIEW_QUESTION')+comparison_count(comparison_records,'SYSTEM_LIMITATION')
        readiness=pd.DataFrame([
            readiness_row('Задание на проектирование',assignment_plan),
            {'Контур':'НТД — ссылки и редакции','Всего':len(normative_rows),'Завершено автоматически':normative_valid_verified,'Подтверждённые несоответствия':0,'Требует внимания':max(0,len(normative_rows)-normative_valid_verified),'Покрытие, %':round(100*normative_valid_verified/max(1,len(normative_rows)),1)},
            readiness_row('НТД — доказательные требования',normative_plan),
            readiness_row('Чек-листы',checklist_plan),
            {'Контур':'Сверка реестров и чертежей','Всего':len(register_comparisons),'Завершено автоматически':_cross_completed(register_comparisons),'Подтверждённые несоответствия':comparison_count(register_comparisons,'PROJECT_FINDING'),'Требует внимания':comparison_count(register_comparisons,'REVIEW_QUESTION')+comparison_count(register_comparisons,'SYSTEM_LIMITATION'),'Покрытие, %':round(100*_cross_completed(register_comparisons)/max(1,len(register_comparisons)),1)},
            {'Контур':'Сверка инженерных параметров','Всего':len(engineering_comparisons),'Завершено автоматически':_cross_completed(engineering_comparisons),'Подтверждённые несоответствия':comparison_count(engineering_comparisons,'PROJECT_FINDING'),'Требует внимания':comparison_count(engineering_comparisons,'REVIEW_QUESTION')+comparison_count(engineering_comparisons,'SYSTEM_LIMITATION'),'Покрытие, %':round(100*_cross_completed(engineering_comparisons)/max(1,len(engineering_comparisons)),1)},
        ])
        sheets.append(('Готовность проверки',_excel_safe_frame(readiness)))
        if not coverage_matrix_df.empty:
            sheets.append(('Карта покрытия',_excel_safe_frame(coverage_matrix_df,columns=[
                'Архетип','Всего','Завершено автоматически','Покрытие, %',
                'Доказательства готовы L3–L5','Доказательное покрытие, %','L4','Основные причины пробелов',
            ])))
        if not limitations_df.empty:
            sheets.append(('Границы автоматизации',_excel_safe_frame(limitations_df.head(40),columns=[
                'Контур','Проверка','Уровень доказательства','Что не получено','Ожидаемые разделы',
            ])))
        if not recommendations_df.empty: sheets.append(('Приоритетные действия', recommendations_df.head(10)))
    elif report_kind == 'gip':
        if not review_plan_df.empty: sheets.append(('Проверка требований', review_plan_df))
        if not coverage_matrix_df.empty: sheets.append(('Карта покрытия',coverage_matrix_df))
        if not limitations_df.empty: sheets.append(('Границы автоматизации',limitations_df))
        if not assignment_atomic_df.empty: sheets.append(('Задание — атомарные условия', assignment_atomic_df))
        if not assignment_gip_df.empty: sheets.append(('Задание на проектирование', assignment_gip_df))
        if not normative_compliance_df.empty: sheets.append(('НТД — требования', normative_compliance_df))
        if not checklist_all_df.empty: sheets.append(('Чек-листы', checklist_all_df))
        if not ai_summary_df.empty: sheets.append(('AI — сводка',ai_summary_df))
        if not ai_execution_df.empty: sheets.append(('AI — решения',ai_execution_df))
        if not recommendations_df.empty: sheets.append(('План действий', recommendations_df))
    if report_kind == 'technical':
        if not coverage_matrix_df.empty: sheets.append(('Карта покрытия',coverage_matrix_df))
        if not limitations_df.empty: sheets.append(('Границы автоматизации',limitations_df))
        if not object_df.empty: sheets.append(('Состав проекта', object_df))
        if not checklist_problem_df.empty: sheets.append(('Чек-листы — диагностика', checklist_problem_df))
        if not normative_df.empty: sheets.append(('Актуальность НТД', normative_df))
        if not normative_compliance_df.empty: sheets.append(('Проверка требований НТД', normative_compliance_df))
        if not normative_requirement_df.empty: sheets.append(('Контекст ссылок НТД', normative_requirement_df))
        if not understanding_df.empty: sheets.append(('Модель проекта', understanding_df))
        if not assignment_df.empty: sheets.append(('Задание — диагностика', assignment_df))
        if not assignment_atomic_df.empty: sheets.append(('Задание — атомарные условия', assignment_atomic_df))
        if not ai_summary_df.empty: sheets.append(('AI — сводка',ai_summary_df))
        if not ai_execution_df.empty: sheets.append(('AI — решения',ai_execution_df))
        if not ai_calls_df.empty: sheets.append(('AI — вызовы',ai_calls_df))
        if not recommendations_df.empty: sheets.append(('План действий', recommendations_df))
        if normative_reference_details:
            detailed_normative_df=pd.DataFrame(normative_reference_details)
            if not detailed_normative_df.empty:
                sheets.append(('НТД — все упоминания', _excel_safe_frame(detailed_normative_df)))
        for sheet_name, frame in _compact_technical_frames(docs, findings, comparison_records, report).items():
            if not frame.empty:
                sheets.append((_safe_sheet_name(sheet_name), frame))


    header_ru = {
      "document":"Документ","document_type":"Тип документа","page":"Страница","parameter_code":"Код показателя",
      "parameter_name":"Показатель","object":"Объект","object_id":"ID объекта","object_hint":"Объект",
      "value":"Значение","value_text":"Исходное значение","unit":"Ед. изм.","status":"Статус",
      "category":"Категория","finding":"Выявленная проблема","recommendation":"Рекомендация",
      "sources":"Источники","source":"Источник","confidence":"Достоверность","reason":"Обоснование",
      "section":"Раздел","section_family":"Раздел","checklist":"Чек-лист","question":"Проверка",
      "result":"Результат","priority":"Приоритет","risk_score":"Оценка риска","risk_level":"Уровень риска",
      "genplan_position":"Позиция по ГП","comparison_scope":"Контур сравнения","binding_key":"Ключ привязки",
      "project":"Проект","project_name":"Проект","analysis_time":"Дата проверки","core_version":"Версия ядра",
      "requirement_id":"ID требования","requirement_text":"Требование","required_value":"Требуемое значение",
      "automatic_section":"Раздел","automatic_checklist":"Чек-лист","execution_class":"Тип проверки",
      "table_title":"Наименование таблицы","table_row":"Строка таблицы","explanation":"Пояснение",
      "reference":"Ссылка/обозначение","canonical_id":"Канонический ID","verified_on":"Дата верификации",
      "verified_revision":"Верифицированная редакция","replacement":"Заменяющий документ",
      "effective_until":"Действует до","official_source":"Официальный источник","official_source_kind":"Тип официального источника",
      "impact_risk":"Оценка влияния","cell_reconstruction":"Способ восстановления требования",
      "evidence_quality_state":"Качество доказательства","requirement_scope":"Область требования",
      "requirement_type":"Тип требования","finding_type":"Тип результата","promotion_method":"Способ подтверждения",
      "semantic_evidence_score":"Смысловая достоверность",
      "check_id":"ID проверки","values_by_section":"Значения по разделам",
      "strong_evidence_count":"Количество сильных доказательств",
      "page_count":"Количество страниц","size_mb":"Размер, МБ",
      "completeness_status":"Статус комплектности","processing_error":"Ошибка обработки",
      "document_profile":"Профиль документа","section_title":"Заголовок раздела",
      "binding_status":"Статус привязки","row_integrity_status":"Целостность строки",
      "row_integrity_reason":"Причина контроля строки","fact_admission_decision":"Допуск факта",
      "fact_admission_score":"Оценка допуска факта","fact_admission_reasons":"Причины допуска факта",
      "evidence_quality_decision":"Решение по качеству доказательства",
      "evidence_trust_grade":"Класс доверия доказательству","physical_trace_level":"Уровень физической трассировки",
      "source_locator":"Адрес источника","project_understanding_binding":"Привязка в модели проекта",
      "comparison_excluded":"Исключено из сравнения","engineering_plausibility_status":"Инженерная правдоподобность",
      "engineering_plausibility_reason":"Причина контроля правдоподобности",
      "engineering_derived_height":"Расчётная высота","possible_decimal_separator_candidate":"Возможная десятичная поправка",
      "match_method":"Метод сопоставления","structural_zone":"Структурная зона",
    }
    normalized_sheets=[]
    for sheet_name,frame in sheets:
        if isinstance(frame,pd.DataFrame):
            frame=frame.rename(columns={c:header_ru.get(str(c),header_label(c)) for c in frame.columns})
            for col in frame.columns:
                if frame[col].dtype == object:
                    frame[col]=frame[col].map(localize_service_value)
        normalized_sheets.append((sheet_name,frame))
    sheets=normalized_sheets

    # XlsxWriter creates a clean OOXML package and avoids repair messages that
    # Excel can show for workbooks containing complex openpyxl metadata.
    with pd.ExcelWriter(
        out,
        engine='xlsxwriter',
        engine_kwargs={'options': {
            'strings_to_formulas': False,
            'strings_to_urls': False,
            'nan_inf_to_errors': False,
        }},
    ) as writer:
        workbook = writer.book
        workbook.set_properties({
            'title': f'ExpertCheck — {project}',
            'subject': 'Автоматизированная проверка проектной документации',
            'author': 'ExpertCheck',
            'company': 'ExpertCheck',
        })
        header_fmt = workbook.add_format({
            'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#1F4E78',
            'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
            'border': 1, 'border_color': '#B8C2CC',
        })
        body_fmt = workbook.add_format({
            'valign': 'top', 'text_wrap': True,
            'bottom': 1, 'bottom_color': '#D9E0E6',
        })
        label_fmt = workbook.add_format({
            'bold': True, 'bg_color': '#D9EAF7', 'valign': 'top',
            'text_wrap': True, 'border': 1, 'border_color': '#D9E0E6',
        })
        value_fmt = workbook.add_format({
            'bg_color': '#F3F6F9', 'valign': 'top', 'text_wrap': True,
            'border': 1, 'border_color': '#D9E0E6',
        })
        bad_fmt = workbook.add_format({'bg_color': '#FCE4D6'})
        warn_fmt = workbook.add_format({'bg_color': '#FFF2CC'})
        ok_fmt = workbook.add_format({'bg_color': '#E2F0D9'})

        used_names: set[str] = set()
        for raw_name, frame in sheets:
            name = _safe_sheet_name(raw_name)
            base = name
            suffix = 2
            while name in used_names:
                tail = f'_{suffix}'
                name = (base[:31-len(tail)] + tail)
                suffix += 1
            used_names.add(name)
            safe_frame = _excel_safe_frame(frame)
            safe_frame.to_excel(writer, sheet_name=name, index=False)
            worksheet = writer.sheets[name]
            worksheet.hide_gridlines(2)
            worksheet.set_row(0, 30, header_fmt)
            if name == 'Резюме':
                worksheet.set_column(0, 0, 34, label_fmt)
                worksheet.set_column(1, 1, 68, value_fmt)
                worksheet.set_landscape()
                worksheet.set_paper(9)
                worksheet.fit_to_pages(1, 0)
                worksheet.set_margins(0.3, 0.3, 0.45, 0.45)
                worksheet.set_footer('&LExpertCheck&CСтраница &P из &N&R&D')
            else:
                worksheet.freeze_panes(1, 0)
                worksheet.set_landscape()
                worksheet.set_paper(9)
                worksheet.fit_to_pages(1, 0)
                worksheet.repeat_rows(0)
                worksheet.set_margins(0.25, 0.25, 0.4, 0.4)
                worksheet.set_header(f'&LExpertCheck&C{name}&R{str(project)[:45].replace("&", "&&")}')
                worksheet.set_footer('&LАвтоматизированная предпроверка&CСтраница &P из &N&R&D')
                if name in {'AI — сводка','AI — вызовы'}:
                    for row_index in range(1, len(safe_frame)+1):
                        worksheet.set_row(row_index, 48)
                if len(safe_frame.columns):
                    worksheet.autofilter(0, 0, max(len(safe_frame), 1), len(safe_frame.columns)-1)
                for col_idx, column in enumerate(safe_frame.columns):
                    values = [str(column)] + [str(v or '') for v in safe_frame[column].head(120)]
                    width = min(max(max((len(v) for v in values), default=0) + 2, 12), 52)
                    worksheet.set_column(col_idx, col_idx, width, body_fmt)
                # Lightweight conditional highlighting without formulas.
                if len(safe_frame) and len(safe_frame.columns):
                    last_row = len(safe_frame)
                    last_col = len(safe_frame.columns)-1
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Высокий', 'format': bad_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Расхождение', 'format': bad_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Требует', 'format': warn_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Средний', 'format': warn_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Совпадает', 'format': ok_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Подтверждено источником', 'format': ok_fmt})
                    worksheet.conditional_format(1, 0, last_row, last_col, {
                        'type': 'text', 'criteria': 'containing', 'value': 'Соответствует', 'format': ok_fmt})

    payload = out.getvalue()
    if not payload.startswith(b'PK'):
        raise ValueError('Не удалось сформировать корректный XLSX-файл.')

    # Full OOXML and workbook round-trip validation before download.
    import zipfile
    from openpyxl import load_workbook
    buffer = io.BytesIO(payload)
    with zipfile.ZipFile(buffer) as archive:
        broken = archive.testzip()
        if broken is not None:
            raise ValueError(f'Повреждён элемент XLSX: {broken}')
    buffer.seek(0)
    check_book = load_workbook(buffer, read_only=True, data_only=False)
    check_book.close()
    return payload


def excel_report(project, version, docs, findings, comparisons):
    """Backward-compatible default: compact GIP report."""
    return structured_excel_report(project, version, docs, findings, comparisons, report_kind='gip')
