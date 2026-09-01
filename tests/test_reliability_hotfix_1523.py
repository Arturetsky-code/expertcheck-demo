from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from core.atomic_requirement_graph import atomize_requirement
from core.atomic_verification_engine import aggregate_atomic_results
from core.project_review_planner import build_review_plan
from core.project_snapshot import build_analysis_snapshot
from core.report_quality_gate import validate_review_plan
from core.semantic_continuation import continue_semantic_analysis
from core.verified_verdict_gate import enforce_verified_verdicts
from studio.data import _stable_report_id, structured_excel_report


def _categorical_atom(state: str = "PASSED") -> dict:
    return {
        "atom_id": "SHIFT-1523-A001",
        "parent_requirement_id": "SHIFT-1523",
        "verification_kind": "VERIFIED_OK",
        "final_verification_kind": "VERIFIED_OK",
        "status": "Соответствует заданию",
        "evidence_level": "L5",
        "adversarial_state": state,
        "adversarial_reasons": [] if state == "PASSED" else ["Недостаточно независимой проверки."],
        "verification_evidence": [{
            "document": "ТХ.pdf", "page": 33, "section": "ТХ",
            "text": "Продолжительность смены 12 часов.",
        }],
        "evidence_candidates": [{
            "document": "ТХ.pdf", "page": 33, "section": "ТХ",
            "text": "Продолжительность смены 12 часов.",
        }],
        "evidence": ["ТХ.pdf, стр. 33: Продолжительность смены 12 часов."],
        "recipe_status": "TRUSTED",
        "proof_kind": "STRUCTURED_VALUE",
    }


def test_parent_aggregation_preserves_l5_gate_and_addressable_evidence():
    parent = {
        "requirement_id": "SHIFT-1523",
        "requirement_text": "Продолжительность смены — 12 часов",
    }
    atoms = [_categorical_atom()]
    assert enforce_verified_verdicts(atoms, domain="assignment")["passed"] == 1
    rows = aggregate_atomic_results([parent], atoms)
    assert enforce_verified_verdicts(rows, domain="assignment")["passed"] == 1
    assert rows[0]["adversarial_state"] == "PASSED"
    assert rows[0]["deep_evidence_state"] == "PASSED"
    assert rows[0]["deep_evidence_candidate_count"] == 1
    plan = build_review_plan(
        assignment_rows=rows, normative_rows=[], checklist_review={"results": []},
    )
    assert validate_review_plan(plan)["status"] == "PASSED"


def test_parent_aggregation_keeps_failed_l5_gate_visible_to_quality_gate():
    parent = {
        "requirement_id": "SHIFT-1523",
        "requirement_text": "Продолжительность смены — 12 часов",
    }
    atoms = [_categorical_atom("BLOCKED")]
    enforce_verified_verdicts(atoms, domain="assignment")
    rows = aggregate_atomic_results([parent], atoms)
    enforce_verified_verdicts(rows, domain="assignment")
    assert rows[0]["verification_kind"] == "REVIEW_QUESTION"
    assert rows[0]["evidence_level"] != "L5"
    plan = build_review_plan(
        assignment_rows=rows, normative_rows=[], checklist_review={"results": []},
    )
    gate = validate_review_plan(plan)
    assert gate["status"] == "PASSED"
    assert not any(item["verification_kind"] in {"VERIFIED_OK", "PROJECT_FINDING"} for item in plan["items"])


def test_report_ids_are_stable_and_never_export_nan():
    row = {"id": float("nan"), "object": "Насосная", "parameter": "Высота", "status": "Требует проверки"}
    first = _stable_report_id("CMP", row)
    second = _stable_report_id("CMP", dict(row))
    assert first == second
    assert first.startswith("CMP-") and "NAN" not in first
    assert _stable_report_id("CMP", {"id": "CMP-777"}) == "CMP-777"


def test_failed_quality_gate_forces_preliminary_xlsx_status_and_stable_question_id():
    question = {
        "plan_id": float("nan"), "domain": "Чек-листы", "title": "Проверить схему",
        "verification_kind": "REVIEW_QUESTION", "status": "Требует проверки",
        "coverage_reason": "Нужна смысловая проверка", "evidence_level": "L4",
    }
    domains = {
        "Задание на проектирование": {"total": 0},
        "НТД": {"total": 0},
        "Чек-листы": {"total": 1, "review": 1, "system_limitation": 0, "coverage_pct": 0},
    }
    docs = [{
        "Файл": "ИОС2.pdf", "Раздел": "ИОС2", "completeness_user_confirmed": True,
        "project_review_plan": {
            "items": [question], "project_findings": 0, "review_questions": 1,
            "system_limitations": 0, "domains": domains,
        },
        "coverage_matrix": {"matrix": [], "coverage_pct": 0, "evidence_coverage_pct": 100},
        "report_quality_gate": {"status": "FAILED", "issues": ["Тестовая причина gate."]},
    }]
    workbook = load_workbook(BytesIO(structured_excel_report(
        "Тест", "15.2.3", docs, [], [], report_kind="gip", checklist_results=[],
    )), read_only=True, data_only=True)
    summary = {row[0].value: row[1].value for row in workbook["Резюме"].iter_rows()}
    assert summary["Статус отчёта"] == "Предварительный — Quality Gate отчёта не пройден"
    question_id = workbook["Вопросы специалисту"]["A2"].value
    assert str(question_id).startswith("Q-") and str(question_id).lower() != "nan"
    workbook.close()


def test_semantic_continuation_rechecks_snapshot_without_source_pdf():
    requirement = {
        "requirement_id": "SHIFT-CONTINUE",
        "requirement_text": "Продолжительность смены — 12 часов",
        "requirement_type": "VALUE_COMPARISON",
        "parameter_code": "SHIFT_DURATION",
        "required_value": 12,
        "unit": "ч",
    }
    atom = next(
        row for row in atomize_requirement(requirement)
        if row.get("atomic_kind") == "VALUE_COMPARISON"
    )
    atom["directed_evidence_candidates"] = [{
        "evidence_state": "verified_candidate", "parameter_code": "SHIFT_DURATION",
        "value": 12, "unit": "ч", "document": "ТХ.pdf", "page": 33,
        "source_trace": "Продолжительность смены 12 часов", "owner_match": True,
    }]
    pages = [{
        "document": "ТХ.pdf", "document_type": "ТХ", "page": 33,
        "text": "Продолжительность смены 12 часов",
    }]
    snapshot = build_analysis_snapshot(
        [{"Файл": "ТХ.pdf"}], page_corpus=pages,
        fact_graph={"facts": []}, object_registry=[],
    )
    document = {
        "Файл": "ТХ.pdf",
        "analysis_snapshot": snapshot,
        "atomic_requirement_graph": {"atoms": [atom], "summary": {"source_requirements": 1}},
        "assignment_compliance": [requirement],
        "automatic_checklist_review": {"results": [], "summary": {}},
        "normative_compliance_audit": [],
        "universal_project_fact_graph": {"facts": []},
        "coverage_acceleration_budget": {
            "assignment_semantic_limit": 10, "checklist_semantic_limit": 10,
        },
    }
    checkpoint = {}
    output = continue_semantic_analysis(
        ([document], [], []), knowledge_root="knowledge",
        semantic_level="off", checkpoint=checkpoint,
    )
    updated = output[0][0]
    assert updated["semantic_continuation"]["source_pdf_required"] is False
    assert updated["assignment_compliance"][0]["adversarial_state"] == "PASSED"
    assert updated["report_quality_gate"]["status"] == "PASSED"
    assert checkpoint["_project_fingerprint"] == snapshot["snapshot_id"]
