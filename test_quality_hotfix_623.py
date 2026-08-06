from io import BytesIO
import zipfile
import pandas as pd
from openpyxl import load_workbook

from core.position_rules import is_date_like_position, normalize_genplan_position
from core.object_register_engine import normalize_position
from studio.data import structured_excel_report


def test_dates_are_not_genplan_positions():
    for value in ('21.05.26', '21.05.2026', '2026.05.21', '21/05/26', '21-05-2026'):
        assert is_date_like_position(value)
        assert normalize_genplan_position(value, allow_integer=True) == ''
    assert normalize_position('21.05.26') == ''
    assert normalize_genplan_position('4.13', allow_integer=True) == '4.13'
    assert normalize_genplan_position('7', allow_integer=True) == '7'


def test_all_reports_are_valid_xlsx_and_nested_values_are_sanitized():
    docs = pd.DataFrame([{
        'document': 'ПЗ\x00.pdf', 'document_type': 'ПЗ', 'page_count': 50,
        'consolidated_registry': [{'Позиция по ГП':'4.13','Наименование объекта':'Здание','Источники':['ПЗ','ПЗУ']}],
        'nested': {'a': [1, 2]},
    }])
    findings = pd.DataFrame([{
        'document':'ПЗ.pdf','document_type':'ПЗ','page':45,'parameter_code':'AREA_BUILD',
        'parameter_name':'Площадь застройки','object_hint':'Здание','value_text':'89,9 м²',
        'context':'текст\x01 с запрещённым символом','nested':{'row':[1,2]},
    }])
    comparisons = pd.DataFrame([{
        'check_id':'XCHK-1','object':'Здание','parameter':'Площадь','status':'Потенциальное расхождение',
        'priority':'Высокий','values_by_section':{'ПЗ':'89,9','ПЗУ':'90,0'},
        'explanation':'=не формула','sources':['ПЗ, стр.45','ПЗУ, стр.12'],
    }])
    risks = [{
        'risk_id':'RISK-1','level':'Высокий','category':'ТЭП','object':'Здание','parameter':'Площадь',
        'finding':'Различаются значения','possible_remark':'Проверить согласованность',
        'recommendation':'Сверить документы','sources':['ПЗ','ПЗУ'],
    }]
    assembly = [{'Ключ':'4.13|здание','Включить':True,'Позиция по ГП':'4.13','Наименование объекта':'Здание','Статус проектирования':'Проектируемый','Основание включения':'ПЗ, стр.45'}]
    checklist = [{'item_no':'1.1','question':'Проверить состав','status':'Нет','evidence':{'pages':[1,2]}}]

    for kind in ('manager','gip','technical'):
        payload = structured_excel_report('Тест', '6.2.3', docs, findings, comparisons,
            report_kind=kind, risks=risks, checklist_results=checklist, assembly_rows_data=assembly)
        assert payload.startswith(b'PK')
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            assert archive.testzip() is None
        workbook = load_workbook(BytesIO(payload), read_only=False, data_only=False)
        assert 'Резюме' in workbook.sheetnames
        assert all(len(name) <= 31 for name in workbook.sheetnames)
        if kind == 'technical':
            assert 'Тех_извлечение' in workbook.sheetnames
            assert len(workbook.sheetnames) <= 12
