from __future__ import annotations

from io import BytesIO
from pathlib import Path
import math
import sys
from openpyxl import load_workbook

ROOT=Path(__file__).parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0,str(ROOT))

from core.atomic_requirement_graph import atomize_requirement
from core.atomic_verification_engine import verify_atomic_requirements, _convert
from core.categorical_consistency import build_categorical_consistency_checks
from core.coverage_matrix import build_coverage_matrix
from core.deep_evidence_intelligence import _apply_final_verdict
from core.external_benchmark import evaluate_benchmark
from core.verification_recipe_compiler_v2 import VerificationRecipeCompilerV2
from studio.data import structured_excel_report


def test_adaptive_drawing_contract_is_executable_but_modality_bound():
    atom = atomize_requirement({
        "requirement_id": "DRAW-1",
        "requirement_text": "План сетей электроснабжения с указанием электрощитового оборудования",
    }, domain="checklist")[0]
    recipe = VerificationRecipeCompilerV2("knowledge").compile(atom)
    assert recipe["recipe_status"] == "TRUSTED"
    assert recipe["pattern_origin"] == "ADAPTIVE_CONTRACT_COMPILER"
    assert recipe["required_modality"] == "DRAWING"

    text_page = [{
        "document": "Раздел ПД №5_ИОС1.1.pdf", "section": "ИОС1", "document_type": "ИОС1", "page": 7,
        "text": "В текстовой части описан план сетей электроснабжения с указанием электрощитового оборудования.",
    }]
    blocked = verify_atomic_requirements([atom], knowledge_root="knowledge", fact_graph={"facts": [], "passages": text_page}, page_corpus=text_page)[0]
    assert blocked["verification_kind"] != "VERIFIED_OK"
    assert blocked["coverage_reason_code"] == "WRONG_EVIDENCE_MODALITY"

    drawing_page = [{
        "document": "Раздел ПД №5_ИОС1.2.pdf", "section": "ИОС1", "document_type": "ИОС1", "page": 3,
        "source_modality": "DRAWING",
        "text": "План сетей электроснабжения с указанием электрощитового оборудования.",
    }]
    verified = verify_atomic_requirements([atom], knowledge_root="knowledge", fact_graph={"facts": [], "passages": drawing_page}, page_corpus=drawing_page)[0]
    assert verified["verification_kind"] == "VERIFIED_OK"
    assert verified["coverage_state"] == "AUTOMATED_COMPLETE"


def test_qualitative_question_does_not_get_a_lexical_shortcut():
    atom = atomize_requirement({
        "requirement_id": "QUAL-1", "requirement_text": "Проверить достаточность и корректность принятых решений",
    }, domain="checklist")[0]
    recipe = VerificationRecipeCompilerV2("knowledge").compile(atom)
    assert not recipe["pattern_id"]
    assert recipe["recipe_status"] == "EXPERIMENTAL"


def test_addressable_opo_classification_conflict_is_a_project_finding():
    pages = [
        {"document": "ПЗ.pdf", "document_type": "ПЗ", "page": 10, "text": "Опасный производственный объект относится к III классу опасности."},
        {"document": "ТХ.pdf", "document_type": "ТХ", "page": 8, "text": "Класс опасности опасного производственного объекта: II."},
    ]
    rows = build_categorical_consistency_checks(pages, [])
    assert len(rows) == 1
    assert rows[0]["finding_type"] == "PROJECT_FINDING"
    assert rows[0]["independent_trusted_sources"] == 2
    assert "стр. 10" in rows[0]["sources"] and "стр. 8" in rows[0]["sources"]


def test_external_benchmark_scores_before_and_after_without_bundling_corpus():
    benchmark = {
        "benchmark_id": "synthetic",
        "cases": [{
            "case_id": "C-1", "category": "IDENTITY_CLASSIFICATION",
            "expected_before": "PROJECT_FINDING", "expected_after": "NO_PROJECT_FINDING",
            "match": {"parameter_codes": ["DANGER_CLASS"], "all_terms": ["класс опасности"]},
        }],
    }
    before = {"comparisons": [{
        "check_code": "CAT-1", "parameter_code": "DANGER_CLASS", "parameter_name": "Класс опасности ОПО",
        "finding_type": "PROJECT_FINDING", "status": "ПОТЕНЦИАЛЬНОЕ РАСХОЖДЕНИЕ",
        "sources": "ПЗ.pdf, стр. 10 | ТХ.pdf, стр. 8",
    }]}
    after = {"comparisons": [{
        "check_code": "CAT-1", "parameter_code": "DANGER_CLASS", "parameter_name": "Класс опасности ОПО",
        "finding_type": "PROJECT_STATUS", "status": "СОВПАДАЕТ",
        "sources": "ПЗ.pdf, стр. 10 | ТХ.pdf, стр. 8",
    }]}
    result = evaluate_benchmark(benchmark, before_payload=before, after_payload=after)
    assert result["summary"]["true_positive"] == 1
    assert result["summary"]["true_negative"] == 1
    assert result["summary"]["precision_pct"] == 100.0
    assert result["summary"]["recall_pct"] == 100.0


def test_density_units_are_compared_in_one_engineering_family():
    assert math.isclose(_convert(1.65, "т/м3", "кг/м3") or 0.0, 1650.0)
    assert math.isclose(_convert(1650.0, "кг/м3", "т/м3") or 0.0, 1.65)


def test_adversarial_downgrade_is_removed_from_completed_coverage():
    row = {
        "verification_kind": "VERIFIED_OK", "status": "Соответствует",
        "coverage_state": "AUTOMATED_COMPLETE", "coverage_reason_code": "EVIDENCE_CONTRACT_SATISFIED",
    }
    _apply_final_verdict(row, {
        "verification_kind": "REVIEW_QUESTION", "verification_state": "Требует проверки специалистом",
        "adversarial_state": "BLOCKED", "adversarial_reasons": ["Недостаточно независимых доказательств"],
    })
    assert row["coverage_state"] == "TARGETED_REVIEW"
    assert row["coverage_reason_code"] == "ADVERSARIAL_OR_SEMANTIC_GATE_BLOCKED"
    assert "INDEPENDENT_SEMANTIC_CONFIRMATION" in row["missing_evidence_slots"]


def test_gip_report_shows_limitations_but_does_not_put_them_in_action_plan():
    plan_item = {
        "plan_id": "CHECK-0001", "domain": "Чек-листы", "title": "Проверка расчёта",
        "status": "SYSTEM_LIMITATION", "verification_kind": "SYSTEM_LIMITATION",
        "coverage_archetype": "CALCULATION_EVIDENCE", "coverage_state": "AUTOMATION_GAP",
        "coverage_reason_code": "NO_ADDRESSABLE_EVIDENCE", "coverage_reason": "Не найден адресный расчёт.",
        "missing_evidence_slots": ["SOURCE_DOCUMENT", "PAGE"], "expected_sections": ["КР"],
    }
    matrix = build_coverage_matrix([plan_item])
    docs = [{
        "Файл": "КР.pdf", "Раздел": "КР", "completeness_user_confirmed": False,
        "project_review_plan": {
            "items": [plan_item], "project_findings": 0, "review_questions": 0, "system_limitations": 1,
            "domains": {
                "Задание на проектирование": {"total": 0}, "НТД": {"total": 0},
                "Чек-листы": {"total": 1, "review": 0, "system_limitation": 1, "coverage_pct": 0},
            },
        },
        "coverage_matrix": matrix, "report_quality_gate": {"status": "PASSED"},
    }]
    payload = structured_excel_report("Тест", "12.0", docs, [], [], report_kind="gip", checklist_results=[])
    book = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert "Проверка требований" in book.sheetnames
    assert "Границы автоматизации" in book.sheetnames
    assert "Карта покрытия" in book.sheetnames
    if "План действий" in book.sheetnames:
        values = " ".join(str(cell.value or "") for row in book["План действий"].iter_rows() for cell in row)
        assert "Не найден адресный расчёт" not in values
    summary_values = " ".join(str(cell.value or "") for row in book["Резюме"].iter_rows() for cell in row)
    assert "Предварительный" in summary_values
    book.close()


if __name__ == "__main__":
    for name, value in sorted(globals().items()):
        if name.startswith("test_") and callable(value):
            value()
