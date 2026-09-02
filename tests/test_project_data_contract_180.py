from __future__ import annotations

import gzip
import io
import json

from openpyxl import load_workbook

from core.project_data_contract import (
    CONTRACT_VERSION,
    enforce_project_data_contract,
)
from core.project_snapshot import load_project_snapshot, project_snapshot_bytes
from studio.data import frames, structured_excel_report
from studio.report_resilience import build_report_isolated


def _document():
    return {
        "Файл": "ТХ.pdf",
        "Тип документа": "ТХ",
        "Страниц": 3,
        "core_version": "18.0-stage1-project-data-contract",
        "assignment_compliance": [],
        "normative_compliance_audit": [],
        "automatic_checklist_review": {"results": []},
        "project_review_plan": {},
        "coverage_matrix": {},
        "semantic_evidence_engine": {},
        "report_quality_gate": {"status": "PASSED", "issues": []},
        "analysis_snapshot": {
            "snapshot_id": "snapshot-18",
            "page_corpus": [{"document": "ТХ.pdf", "page": 1, "text": "Производительность 100 т/ч"}],
        },
    }


def test_contract_repairs_mixed_public_result_values():
    documents, findings, comparisons, audit = enforce_project_data_contract(
        [_document()],
        [{
            "document": "ТХ.pdf", "page": 1, "parameter_code": "CAPACITY",
            "object_hint": 17, "value": float("nan"), "unit": "т/ч",
            "evidence": "ТХ.pdf, стр. 1",
        }],
        [{
            "object": 17, "parameter_name": "Производительность", "status": 422,
            "cross_section_gate_reasons": ["Нет контроля", None, {"code": "MISSING_CONTROL"}],
            "dependency_diagnostics": ["invalid"],
        }],
        source="test",
    )

    assert audit["version"] == CONTRACT_VERSION
    assert audit["status"] == "REPAIRED"
    assert not audit["fatal_issues"]
    assert documents[0]["contract_id"].startswith("DOC-")
    assert findings[0]["finding_id"].startswith("FND-")
    assert findings[0]["value"] is None
    assert findings[0]["object_hint"] == "17"
    assert findings[0]["evidence"] == ["ТХ.pdf, стр. 1"]
    assert comparisons[0]["comparison_id"].startswith("CMP-")
    assert comparisons[0]["status"] == "422"
    assert comparisons[0]["dependency_diagnostics"] == {}


def test_contract_fingerprint_is_repeatable():
    args = (
        [_document()],
        [{"document": "ТХ.pdf", "page": 1, "parameter_code": "CAPACITY", "value": 100, "unit": "т/ч"}],
        [{"object": "Комплекс", "parameter_name": "Производительность", "status": "СОВПАДАЕТ"}],
    )
    first = enforce_project_data_contract(*args, source="first")[3]
    second = enforce_project_data_contract(*args, source="second")[3]
    assert first["result_identity_fingerprint"] == second["result_identity_fingerprint"]


def test_contract_rejects_non_mapping_result_rows():
    documents, findings, comparisons, audit = enforce_project_data_contract(
        [_document()],
        ["неструктурированный результат"],
        [],
        source="test_invalid_row",
    )

    assert len(documents) == 1
    assert findings == []
    assert comparisons == []
    assert audit["status"] == "FAILED"
    assert any("findings[0]" in issue for issue in audit["fatal_issues"])


def test_ui_frames_migrates_legacy_result_once_and_persists_contract():
    result = (
        [_document()],
        [{"document": "ТХ.pdf", "page": 1, "object_hint": 17, "value": float("nan")}],
        [{"object": 17, "parameter_name": "Производительность", "status": 422}],
    )
    documents, findings, comparisons = frames(result)
    assert documents.iloc[0]["project_data_contract"]["version"] == CONTRACT_VERSION
    assert findings.iloc[0]["value"] is None
    assert comparisons.iloc[0]["status"] == "422"
    assert result[0][0]["project_data_contract"]["source"] == "ui_result_boundary"

    # The second UI rerun uses the persisted contract instead of re-migrating.
    second_documents, _, _ = frames(result)
    assert second_documents.iloc[0]["project_data_contract"]["source"] == "ui_result_boundary"


def test_snapshot_roundtrip_enforces_contract_without_source_pdf():
    payload = project_snapshot_bytes(
        [_document()],
        [{"document": "ТХ.pdf", "page": 1, "object_hint": 17, "value": float("nan")}],
        [{"object": 17, "parameter_name": "Производительность", "status": 422}],
    )
    raw = json.loads(gzip.decompress(payload).decode("utf-8"))
    assert raw["snapshot_export_contract"]["version"] == CONTRACT_VERSION

    loaded = load_project_snapshot(payload)
    assert loaded["snapshot_load_contract"]["status"] in {"PASSED", "REPAIRED"}
    assert loaded["findings"][0]["value"] is None
    assert loaded["comparisons"][0]["status"] == "422"


def test_technical_report_exposes_data_contract_audit():
    payload = structured_excel_report(
        "Проект", "ExpertCheck 18.0 Development · Verified Platform",
        [_document()],
        [{"document": "ТХ.pdf", "page": 1, "object_hint": 17, "value": float("nan")}],
        [{"object": 17, "parameter_name": "Производительность", "status": 422}],
        report_kind="technical", checklist_results=[],
    )
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    assert "Контроль данных" in workbook.sheetnames
    rows = dict(workbook["Контроль данных"].iter_rows(min_row=2, values_only=True))
    assert rows["Версия контракта"] == CONTRACT_VERSION
    assert rows["Статус"] == "REPAIRED"
    assert rows["Исправлено значений"] > 0


def test_individual_report_failure_is_isolated():
    failures: list[Exception] = []

    def broken_export():
        raise TypeError("повреждена только одна выгрузка")

    payload = build_report_isolated(broken_export, failures.append)

    assert payload is None
    assert len(failures) == 1
    assert isinstance(failures[0], TypeError)


def test_ten_repeated_xlsx_and_snapshot_exports_are_stable():
    fingerprints: set[str] = set()
    for _run in range(10):
        snapshot = load_project_snapshot(project_snapshot_bytes(
            [_document()],
            [{"document": "ТХ.pdf", "page": 1, "parameter_code": "CAPACITY", "value": 100}],
            [{"object": "Комплекс", "parameter_name": "Производительность", "status": "СОВПАДАЕТ"}],
        ))
        fingerprints.add(snapshot["snapshot_load_contract"]["result_identity_fingerprint"])

        for report_kind in ("manager", "gip", "technical"):
            payload = structured_excel_report(
                "Проект", "ExpertCheck 18.0 Development · Verified Platform",
                [_document()],
                [{"document": "ТХ.pdf", "page": 1, "parameter_code": "CAPACITY", "value": 100}],
                [{"object": "Комплекс", "parameter_name": "Производительность", "status": "СОВПАДАЕТ"}],
                report_kind=report_kind, checklist_results=[],
            )
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            assert workbook.sheetnames
            if report_kind == "technical":
                assert "Контроль данных" in workbook.sheetnames
            workbook.close()

    assert len(fingerprints) == 1
